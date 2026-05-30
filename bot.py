"""
Telegram-бот для учёта USDT-сделок, кэш-движений и арбитражных сделок.

УМЕЕТ ПОНИМАТЬ:
  USDT-сделки:
    Продал Гиге 475600/76.262
    Купил у Стефа 1020*74
    Влад продал Клиенту 96000/75.7

  Кэш-движения (только рубли):
    Выдал Владу 72 600
    Принял от Адахана 729 100
    + 96000 от клиента
    - 72600 Владу

  Арбитраж (сделка без своих средств):
    Сделка без моих средств
    Купил у Германа 196500/75=2620
    Продал Стефу 196500/75.5=2602,649
  → Профит +17.35 USDT, касса не меняется.

ПОДТВЕРЖДЕНИЕ:
  После сделки бот показывает кнопки [✅ Записать] [❌ Отмена].
  По умолчанию подтверждаем USDT-сделки и арбитраж.
  Кэш-движения по умолчанию записываются сразу (можно изменить через /confirm).

ТИПЫ ЧАТОВ (/settype):
  main   — личный чат: все операции
  field  — чат сотрудника (Влада): сделки в поле
  common — общая касса: приходы/расходы
"""

import asyncio
import logging
import os
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
)

# ────────────────── НАСТРОЙКИ ──────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()

# Где хранить базу. Если задана переменная окружения DATA_DIR (на Railway —
# точка монтирования Volume, например /data), база живёт там и НЕ стирается
# при передеплоях. Если переменной нет — база лежит рядом с bot.py (как раньше).
_data_dir = os.getenv("DATA_DIR", "").strip()
if _data_dir:
    DATA_DIR = Path(_data_dir)
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
    except OSError:
        DATA_DIR = Path(__file__).parent
else:
    DATA_DIR = Path(__file__).parent
DB_PATH = DATA_DIR / "trades.db"

# Часовой пояс пользователя для отчётов (Красноярск = UTC+7).
# Можно переопределить переменной окружения TZ_OFFSET (в часах).
TZ_OFFSET_HOURS = int(os.getenv("TZ_OFFSET", "7"))
LOCAL_TZ = timezone(timedelta(hours=TZ_OFFSET_HOURS))

if not BOT_TOKEN:
    raise SystemExit(
        "❌ Не задан BOT_TOKEN.\n"
        "Создай файл .env рядом с bot.py и положи туда: BOT_TOKEN=твой_токен"
    )

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger(__name__)


# ────────────────── БАЗА ДАННЫХ ──────────────────
@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # SQLite не умеет LOWER() для не-ASCII, регистрируем Python-функцию
    conn.create_function("pylower", 1, lambda s: s.lower() if s else "")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    """Создаёт схему и мигрирует старые сделки в новую таблицу transactions."""
    with db() as conn:
        version = conn.execute("PRAGMA user_version").fetchone()[0]

        # Старая схема — оставляем для совместимости со старыми данными
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS trades (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id       INTEGER NOT NULL,
                ts            TEXT    NOT NULL,
                kind          TEXT    NOT NULL,
                counterparty  TEXT,
                usdt          REAL    NOT NULL,
                rate          REAL    NOT NULL,
                rubles        REAL    NOT NULL,
                raw_text      TEXT
            );
            CREATE TABLE IF NOT EXISTS state (
                chat_id        INTEGER PRIMARY KEY,
                ruble_balance  REAL DEFAULT 0,
                usdt_balance   REAL DEFAULT 0,
                extra_rubles   REAL DEFAULT 0
            );
            CREATE INDEX IF NOT EXISTS idx_trades_chat ON trades(chat_id, ts);
        """)

        conn.executescript("""
            CREATE TABLE IF NOT EXISTS transactions (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id        INTEGER NOT NULL,
                ts             TEXT    NOT NULL,
                type           TEXT    NOT NULL,
                counterparty   TEXT,
                doer           TEXT,
                usdt           REAL    DEFAULT 0,
                rate           REAL    DEFAULT 0,
                rubles         REAL    NOT NULL,
                raw_text       TEXT,
                photo_file_id  TEXT,
                note           TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_tx_chat ON transactions(chat_id, id);
            CREATE INDEX IF NOT EXISTS idx_tx_party ON transactions(counterparty);
        """)

        # v2: миграция из старой trades
        if version < 2:
            old_count = conn.execute("SELECT COUNT(*) AS c FROM trades").fetchone()["c"]
            if old_count > 0:
                rows = conn.execute("SELECT * FROM trades ORDER BY id").fetchall()
                for r in rows:
                    conn.execute(
                        "INSERT INTO transactions (chat_id, ts, type, counterparty, usdt, rate, rubles, raw_text) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                        (r["chat_id"], r["ts"], r["kind"], r["counterparty"],
                         r["usdt"], r["rate"], r["rubles"], r["raw_text"])
                    )
                log.info(f"Перенёс {old_count} старых сделок в transactions.")
            conn.execute("PRAGMA user_version = 2")

        # v3: добавляем status, batch_id, и поля для арбитража
        if version < 3:
            cols = {r["name"] for r in conn.execute("PRAGMA table_info(transactions)").fetchall()}
            if "status" not in cols:
                conn.execute("ALTER TABLE transactions ADD COLUMN status TEXT DEFAULT 'confirmed'")
            if "batch_id" not in cols:
                conn.execute("ALTER TABLE transactions ADD COLUMN batch_id INTEGER")
            # Поля для арбитража: что получили / что отдали в USDT
            if "usdt_in" not in cols:
                conn.execute("ALTER TABLE transactions ADD COLUMN usdt_in REAL DEFAULT 0")
            if "usdt_out" not in cols:
                conn.execute("ALTER TABLE transactions ADD COLUMN usdt_out REAL DEFAULT 0")
            if "partner_in" not in cols:
                conn.execute("ALTER TABLE transactions ADD COLUMN partner_in TEXT")
            if "partner_out" not in cols:
                conn.execute("ALTER TABLE transactions ADD COLUMN partner_out TEXT")

            # Все старые транзакции — confirmed
            conn.execute("UPDATE transactions SET status = 'confirmed' WHERE status IS NULL")
            conn.execute("PRAGMA user_version = 3")
            log.info("Применил миграцию v3 (status, batch_id, поля арбитража).")

        # v4: типы долгов добавлены в код (loan_out, loan_in, debt_repay_in, debt_repay_out),
        # отдельной схемы не нужно — используется та же transactions с новыми type.
        if version < 4:
            conn.execute("PRAGMA user_version = 4")
            log.info("Применил миграцию v4 (типы долгов).")

        # Настройки чата (тип, режим подтверждения)
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS chat_settings (
                chat_id       INTEGER PRIMARY KEY,
                chat_type     TEXT DEFAULT 'main',     -- 'main' | 'field' | 'common'
                confirm_mode  TEXT DEFAULT 'trades'    -- 'all' | 'trades' | 'big' | 'off'
            );
        """)

        # v5: колонки для напоминаний (бэкап, оплата Railway)
        if version < 5:
            existing = {r[1] for r in conn.execute("PRAGMA table_info(chat_settings)").fetchall()}
            to_add = {
                "reminders_on":        "INTEGER DEFAULT 0",   # 0/1 — включены ли напоминания
                "last_backup_ts":      "TEXT",                # когда последний раз делали /backup
                "last_backup_remind":  "TEXT",                # дата последнего напоминания о бэкапе (ГГГГ-ММ-ДД)
                "pay_date":            "TEXT",                # дата оплаты Railway (ГГГГ-ММ-ДД)
                "last_pay_remind":     "TEXT",                # дата последнего напоминания об оплате
            }
            for col, decl in to_add.items():
                if col not in existing:
                    conn.execute(f"ALTER TABLE chat_settings ADD COLUMN {col} {decl}")
            conn.execute("PRAGMA user_version = 5")
            log.info("Применил миграцию v5 (напоминания).")

        # v6: отдельный счётчик дохода с арбитража (не смешивается с кошельком)
        if version < 6:
            existing = {r[1] for r in conn.execute("PRAGMA table_info(state)").fetchall()}
            if "arb_profit_usdt" not in existing:
                conn.execute("ALTER TABLE state ADD COLUMN arb_profit_usdt REAL DEFAULT 0")
            conn.execute("PRAGMA user_version = 6")
            log.info("Применил миграцию v6 (отдельный счётчик арбитража).")


# ───── Состояние кассы/кошелька (только для confirmed транзакций) ─────
def get_state(chat_id: int) -> dict:
    with db() as conn:
        row = conn.execute("SELECT * FROM state WHERE chat_id = ?", (chat_id,)).fetchone()
        if not row:
            conn.execute("INSERT INTO state (chat_id) VALUES (?)", (chat_id,))
            return {"ruble_balance": 0.0, "usdt_balance": 0.0,
                    "extra_rubles": 0.0, "arb_profit_usdt": 0.0}
        d = dict(row)
        d.setdefault("arb_profit_usdt", 0.0)
        return d


def update_state(chat_id: int, d_rubles: float = 0, d_usdt: float = 0,
                 d_arb_profit: float = 0):
    """
    Обновить состояние. d_arb_profit меняет ТОЛЬКО отдельный счётчик
    дохода с арбитража, не трогая кошелёк USDT и кассу.
    """
    with db() as conn:
        conn.execute("INSERT OR IGNORE INTO state (chat_id) VALUES (?)", (chat_id,))
        conn.execute(
            "UPDATE state SET ruble_balance = ruble_balance + ?, "
            "usdt_balance = usdt_balance + ?, "
            "arb_profit_usdt = COALESCE(arb_profit_usdt, 0) + ? "
            "WHERE chat_id = ?",
            (d_rubles, d_usdt, d_arb_profit, chat_id)
        )


# ───── Настройки чата ─────
def get_chat_settings(chat_id: int) -> dict:
    with db() as conn:
        row = conn.execute("SELECT * FROM chat_settings WHERE chat_id = ?", (chat_id,)).fetchone()
        if not row:
            conn.execute("INSERT INTO chat_settings (chat_id) VALUES (?)", (chat_id,))
            return {"chat_id": chat_id, "chat_type": "main", "confirm_mode": "trades",
                    "reminders_on": 0, "last_backup_ts": None, "last_backup_remind": None,
                    "pay_date": None, "last_pay_remind": None}
        return dict(row)


def set_chat_setting(chat_id: int, **kw):
    with db() as conn:
        conn.execute("INSERT OR IGNORE INTO chat_settings (chat_id) VALUES (?)", (chat_id,))
        allowed = ("chat_type", "confirm_mode", "reminders_on",
                   "last_backup_ts", "last_backup_remind", "pay_date", "last_pay_remind")
        for k, v in kw.items():
            if k in allowed:
                conn.execute(f"UPDATE chat_settings SET {k} = ? WHERE chat_id = ?", (v, chat_id))


# ───── Сохранение транзакций ─────
def save_tx(chat_id, type_, counterparty, rubles, *, usdt=0, rate=0, raw="",
            doer=None, status="confirmed", batch_id=None,
            usdt_in=0, usdt_out=0, partner_in=None, partner_out=None) -> int:
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO transactions "
            "(chat_id, ts, type, counterparty, doer, usdt, rate, rubles, raw_text, "
            " status, batch_id, usdt_in, usdt_out, partner_in, partner_out) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (chat_id, datetime.utcnow().isoformat(), type_, counterparty, doer,
             usdt, rate, rubles, raw, status, batch_id,
             usdt_in, usdt_out, partner_in, partner_out)
        )
        return cur.lastrowid


def get_avg_rates_excluding(chat_id: int, exclude_id: int, period: str = "cycle") -> dict:
    """
    То же что get_avg_rates, но исключает транзакцию с указанным id.
    Используется чтобы посчитать средние ДО только что записанной сделки —
    для подсветки невыгодных сделок («продал по 73, а средняя закупка 74.5»).
    """
    where = "chat_id = ? AND status = 'confirmed' AND id != ?"
    params = [chat_id, exclude_id]

    if period == "cycle":
        start_id = get_cycle_start_id(chat_id)
        if start_id > 0 and start_id <= exclude_id:
            # Цикл может включать только что добавленную операцию — используем её как старт,
            # но не уходим назад если она открыла цикл сама
            where += " AND id >= ?"
            params.append(start_id)

    with db() as conn:
        sell = conn.execute(
            f"SELECT COALESCE(SUM(rubles),0) r, COALESCE(SUM(usdt),0) u "
            f"FROM transactions WHERE {where} AND type = 'sell'",
            params).fetchone()
        buy = conn.execute(
            f"SELECT COALESCE(SUM(rubles),0) r, COALESCE(SUM(usdt),0) u "
            f"FROM transactions WHERE {where} AND type = 'buy'",
            params).fetchone()
    return {
        "avg_sell": (sell["r"] / sell["u"]) if sell["u"] else None,
        "avg_buy":  (buy["r"]  / buy["u"])  if buy["u"]  else None,
    }


def get_cycle_start_id(chat_id: int, cycle_zero_threshold: float = 10.0) -> int:
    """
    Возвращает ID первой транзакции ТЕКУЩЕГО цикла.

    Цикл = период между моментами, когда USDT-кошелёк был ≤ порога (по умолчанию 10).
    Алгоритм:
      1. Берём ВСЕ confirmed транзакции в хронологическом порядке.
      2. Симулируем USDT-баланс с нуля.
      3. Запоминаем ID последней операции, ПОСЛЕ которой баланс упал до ≤ порога.
      4. Текущий цикл = всё что идёт ПОСЛЕ этой операции.
      5. Если кошелёк никогда не обнулялся — текущий цикл = весь учёт.

    Возвращает 0 если транзакций нет или цикл начался с самого начала.
    """
    with db() as conn:
        rows = conn.execute(
            "SELECT id, type, usdt FROM transactions "
            "WHERE chat_id = ? AND status = 'confirmed' "
            "AND type IN ('sell', 'buy', 'arb') "
            "ORDER BY id",
            (chat_id,)).fetchall()

    balance = 0.0
    last_zero_id = 0
    for r in rows:
        t = r["type"]
        u = r["usdt"] or 0
        if t == "buy":
            balance += u
        elif t == "sell":
            balance -= u
        elif t == "arb":
            balance += u  # арбитраж даёт профит в USDT
        if balance <= cycle_zero_threshold:
            last_zero_id = r["id"]

    # Текущий цикл начинается с ПЕРВОЙ транзакции после last_zero_id
    with db() as conn:
        nxt = conn.execute(
            "SELECT MIN(id) AS mid FROM transactions "
            "WHERE chat_id = ? AND id > ? AND status = 'confirmed' "
            "AND type IN ('sell', 'buy', 'arb')",
            (chat_id, last_zero_id)).fetchone()
    return nxt["mid"] or 0


def get_avg_rates(chat_id: int, period: str = "all", days: int = 30) -> dict:
    """
    Средневзвешенные курсы — только по confirmed транзакциям.

    period:
      'all'   — всё время (по умолчанию)
      'cycle' — текущий цикл (с момента когда кошелёк был ≤ 10 USDT)
      'days'  — последние `days` дней (по умолчанию 30)
    """
    # WHERE-условие и параметры
    where = "chat_id = ? AND status = 'confirmed'"
    params = [chat_id]

    if period == "cycle":
        start_id = get_cycle_start_id(chat_id)
        if start_id > 0:
            where += " AND id >= ?"
            params.append(start_id)
        # если start_id == 0 — цикл = всё время (нет старых обнулений)
    elif period == "days":
        cutoff = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
        where += " AND ts >= ?"
        params.append(cutoff)
    # period == 'all' — без фильтров

    with db() as conn:
        sell = conn.execute(
            f"SELECT COALESCE(SUM(rubles),0) r, COALESCE(SUM(usdt),0) u "
            f"FROM transactions WHERE {where} AND type = 'sell'",
            params).fetchone()
        buy = conn.execute(
            f"SELECT COALESCE(SUM(rubles),0) r, COALESCE(SUM(usdt),0) u "
            f"FROM transactions WHERE {where} AND type = 'buy'",
            params).fetchone()
        arb_profit = conn.execute(
            f"SELECT COALESCE(SUM(usdt),0) p FROM transactions "
            f"WHERE {where} AND type = 'arb'",
            params).fetchone()
        tx_count = conn.execute(
            f"SELECT COUNT(*) AS c FROM transactions WHERE {where} "
            f"AND type IN ('sell','buy','arb')",
            params).fetchone()

    avg_sell = (sell["r"] / sell["u"]) if sell["u"] else None
    avg_buy  = (buy["r"]  / buy["u"])  if buy["u"]  else None

    # Реализованная прибыль = (продано USDT) × (средняя продажа − средняя закупка) + арбитраж × средняя продажа
    realized_profit_rub = None
    if avg_sell is not None and avg_buy is not None and sell["u"] > 0:
        # Сколько USDT реально «прошло» через цикл купи-продай
        traded_usdt = min(sell["u"], buy["u"])
        realized_profit_rub = traded_usdt * (avg_sell - avg_buy)
        # Арбитражный профит в USDT × текущая средняя цена продажи
        if arb_profit["p"] > 0:
            realized_profit_rub += arb_profit["p"] * avg_sell

    return {
        "period":            period,
        "days":              days if period == "days" else None,
        "avg_sell":          avg_sell,
        "avg_buy":           avg_buy,
        "spread":            (avg_sell - avg_buy) if (avg_sell and avg_buy) else None,
        "total_sold_usdt":   sell["u"],
        "total_bought_usdt": buy["u"],
        "total_received":    sell["r"],
        "total_spent":       buy["r"],
        "arb_profit_usdt":   arb_profit["p"],
        "tx_count":          tx_count["c"],
        "realized_profit_rub": realized_profit_rub,
    }


# ───── Долги ─────
def name_root(name: str) -> str:
    """
    Возвращает нормализованный «корень» имени для группировки долгов.
    Русские имена меняются по падежам («Михаил» → «Михаилу» → «Михаила»),
    бот считал бы их разными людьми. Решение — брать первые 3 буквы.

    Стратегия:
      • Латиница / цифры — возвращаем как есть в lowercase (MTX, Mtex).
      • Русское слово — первые 3 буквы в lowercase.
        («Михаил» / «Михаила» / «Михаилу» → "мих";
         «Вася»   / «Васи»    / «Васе»    → "вас";
         «Анна»   / «Анной»   / «Анне»    → "анн";
         «Юра»    / «Юры»     / «Юре»     → "юра".)

    Компромисс: разные имена с одинаковыми первыми 3 буквами схлопнутся
    («Михаил» и «Михей» оба → «мих», «Иван» и «Ивановский» оба → «ива»).
    Если такое случается на практике — увидишь в /debts и переименуешь одного
    контрагента, например на «Михей-Москва» — корень станет другим.
    """
    if not name:
        return ""
    first_word = re.split(r"[\s\(\),]+", name.strip())[0]
    if not first_word:
        return ""
    lower = first_word.lower()

    # Латиница/смешанное — не трогаем
    if not any("а" <= ch <= "я" or ch == "ё" for ch in lower):
        return lower

    # Русское — первые 3 буквы (или всё слово если оно короче)
    return lower[:3]


def get_debts(chat_id: int) -> dict:
    """
    Считает открытые долги по этому чату.
    Имена группируются через name_root, чтобы «Михаил» и «Михаилу» считались
    одним человеком.

    Возвращает {
        "owed_to_us": [(name, amount), ...],
        "we_owe":     [(name, amount), ...],
        "total_owed_to_us": float,
        "total_we_owe": float,
    }
    """
    with db() as conn:
        # Все долговые транзакции
        all_loan_rows = conn.execute(
            "SELECT id, type, counterparty, rubles FROM transactions "
            "WHERE chat_id = ? AND status = 'confirmed' "
            "AND type IN ('loan_out', 'loan_in', 'debt_repay_in', 'debt_repay_out') "
            "ORDER BY id",
            (chat_id,)).fetchall()

    # Агрегируем по name_root
    owed_acc = {}  # root → {"bal": float, "display_name": str}
    we_owe_acc = {}
    for r in all_loan_rows:
        cp = r["counterparty"] or "—"
        root = name_root(cp)
        if r["type"] == "loan_out":
            d = owed_acc.setdefault(root, {"bal": 0.0, "display_name": cp})
            d["bal"] += r["rubles"]; d["display_name"] = cp
        elif r["type"] == "debt_repay_in":
            d = owed_acc.setdefault(root, {"bal": 0.0, "display_name": cp})
            d["bal"] -= r["rubles"]
        elif r["type"] == "loan_in":
            d = we_owe_acc.setdefault(root, {"bal": 0.0, "display_name": cp})
            d["bal"] += r["rubles"]; d["display_name"] = cp
        elif r["type"] == "debt_repay_out":
            d = we_owe_acc.setdefault(root, {"bal": 0.0, "display_name": cp})
            d["bal"] -= r["rubles"]

    owed_to_us = [(v["display_name"], round(v["bal"], 2))
                  for v in owed_acc.values() if v["bal"] > 0.01]
    we_owe = [(v["display_name"], round(v["bal"], 2))
              for v in we_owe_acc.values() if v["bal"] > 0.01]

    owed_to_us.sort(key=lambda x: -x[1])
    we_owe.sort(key=lambda x: -x[1])

    return {
        "owed_to_us":         owed_to_us,
        "we_owe":             we_owe,
        "total_owed_to_us":   sum(a for _, a in owed_to_us),
        "total_we_owe":       sum(a for _, a in we_owe),
    }


def get_debt_history(chat_id: int, name: str) -> list:
    """История долговых операций по корню имени (регистронезависимо, через падежи)."""
    target_root = name_root(name)
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
            "AND type IN ('loan_out', 'loan_in', 'debt_repay_in', 'debt_repay_out') "
            "ORDER BY id",
            (chat_id,)).fetchall()
    # Фильтруем по корню
    return [r for r in rows if name_root(r["counterparty"] or "") == target_root]


# ───── Клиенты ─────
def get_client_stats(chat_id: int, name: str) -> dict:
    """
    Полная статистика по одному клиенту (по корню имени — учитывает падежи и @username).

    Возвращает dict с агрегатами по сделкам, кэшу и долгам, либо None если
    операций с этим контрагентом нет.
    """
    target_root = name_root(name)
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
            "ORDER BY id",
            (chat_id,)).fetchall()

    mine = [r for r in rows
            if name_root(r["counterparty"] or "") == target_root
            or name_root(r["partner_in"] or "") == target_root
            or name_root(r["partner_out"] or "") == target_root]
    if not mine:
        return None

    # Отображаемое имя — самое частое написание
    name_counts = {}
    for r in mine:
        cp = r["counterparty"]
        if cp and name_root(cp) == target_root:
            name_counts[cp] = name_counts.get(cp, 0) + 1
    display_name = max(name_counts, key=name_counts.get) if name_counts else name

    sell_usdt = sell_rub = 0.0
    buy_usdt = buy_rub = 0.0
    sell_count = buy_count = 0
    cash_in = cash_out = 0.0
    loan_out = loan_in = repay_in = repay_out = 0.0
    arb_count = 0
    first_id = mine[0]["id"]
    last_id = mine[-1]["id"]
    first_ts = mine[0]["ts"]
    last_ts = mine[-1]["ts"]

    for r in mine:
        t = r["type"]
        if t == "sell":
            sell_usdt += r["usdt"] or 0; sell_rub += r["rubles"] or 0; sell_count += 1
        elif t == "buy":
            buy_usdt += r["usdt"] or 0; buy_rub += r["rubles"] or 0; buy_count += 1
        elif t == "cash_in":
            cash_in += r["rubles"] or 0
        elif t == "cash_out":
            cash_out += r["rubles"] or 0
        elif t == "loan_out":
            loan_out += r["rubles"] or 0
        elif t == "loan_in":
            loan_in += r["rubles"] or 0
        elif t == "debt_repay_in":
            repay_in += r["rubles"] or 0
        elif t == "debt_repay_out":
            repay_out += r["rubles"] or 0
        elif t == "arb":
            arb_count += 1

    avg_sell = (sell_rub / sell_usdt) if sell_usdt else None
    avg_buy = (buy_rub / buy_usdt) if buy_usdt else None
    # Спред с клиентом: если он у нас покупает (наш sell) и продаёт нам (наш buy)
    spread = (avg_sell - avg_buy) if (avg_sell and avg_buy) else None

    debt_balance = (loan_out - repay_in) - (loan_in - repay_out)  # >0 он нам должен

    return {
        "display_name":   display_name,
        "root":           target_root,
        "total_ops":      len(mine),
        "sell_count":     sell_count,
        "buy_count":      buy_count,
        "sell_usdt":      sell_usdt,
        "buy_usdt":       buy_usdt,
        "sell_rub":       sell_rub,
        "buy_rub":        buy_rub,
        "avg_sell":       avg_sell,
        "avg_buy":        avg_buy,
        "spread":         spread,
        "turnover_rub":   sell_rub + buy_rub,
        "turnover_usdt":  sell_usdt + buy_usdt,
        "cash_in":        cash_in,
        "cash_out":       cash_out,
        "loan_out":       loan_out,
        "loan_in":        loan_in,
        "repay_in":       repay_in,
        "repay_out":      repay_out,
        "debt_balance":   debt_balance,
        "arb_count":      arb_count,
        "first_id":       first_id,
        "last_id":        last_id,
        "first_ts":       first_ts,
        "last_ts":        last_ts,
        "rows":           mine,
    }


def get_all_clients(chat_id: int) -> list:
    """
    Список всех клиентов с агрегатами по обороту USDT-сделок.
    Возвращает список dict, отсортированный по обороту (рубли) убыв.
    Учитывает только тех, с кем были USDT-сделки (sell/buy) — чистые
    кэш-контрагенты и должники не считаются «клиентами по USDT».
    """
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
            "AND type IN ('sell', 'buy') ORDER BY id",
            (chat_id,)).fetchall()

    clients = {}  # root → агрегат
    for r in rows:
        cp = r["counterparty"] or "—"
        root = name_root(cp)
        c = clients.setdefault(root, {
            "display_name": cp, "name_counts": {},
            "turnover_rub": 0.0, "turnover_usdt": 0.0,
            "sell_count": 0, "buy_count": 0,
            "sell_usdt": 0.0, "sell_rub": 0.0,
            "buy_usdt": 0.0, "buy_rub": 0.0,
            "last_id": 0,
        })
        c["name_counts"][cp] = c["name_counts"].get(cp, 0) + 1
        c["turnover_rub"] += r["rubles"] or 0
        c["turnover_usdt"] += r["usdt"] or 0
        c["last_id"] = max(c["last_id"], r["id"])
        if r["type"] == "sell":
            c["sell_count"] += 1; c["sell_usdt"] += r["usdt"] or 0; c["sell_rub"] += r["rubles"] or 0
        else:
            c["buy_count"] += 1; c["buy_usdt"] += r["usdt"] or 0; c["buy_rub"] += r["rubles"] or 0

    result = []
    for root, c in clients.items():
        display = max(c["name_counts"], key=c["name_counts"].get)
        avg_sell = (c["sell_rub"] / c["sell_usdt"]) if c["sell_usdt"] else None
        avg_buy = (c["buy_rub"] / c["buy_usdt"]) if c["buy_usdt"] else None
        spread = (avg_sell - avg_buy) if (avg_sell and avg_buy) else None
        result.append({
            "root": root, "display_name": display,
            "turnover_rub": c["turnover_rub"], "turnover_usdt": c["turnover_usdt"],
            "ops": c["sell_count"] + c["buy_count"],
            "sell_count": c["sell_count"], "buy_count": c["buy_count"],
            "avg_sell": avg_sell, "avg_buy": avg_buy, "spread": spread,
            "last_id": c["last_id"],
        })
    result.sort(key=lambda x: -x["turnover_rub"])
    return result


# ───── Отчёты за период ─────
def _period_bounds(period: str):
    """
    Возвращает (start_utc_iso, end_utc_iso, title) для периода в ЛОКАЛЬНОМ времени
    пользователя (LOCAL_TZ), сконвертированные в UTC-строки для сравнения с ts.

    period: 'day' | 'week' | 'month'
    """
    now_local = datetime.now(LOCAL_TZ)
    if period == "day":
        start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
        title = "Сегодня (" + start_local.strftime("%d.%m") + ")"
    elif period == "week":
        # Неделя с понедельника
        monday = now_local - timedelta(days=now_local.weekday())
        start_local = monday.replace(hour=0, minute=0, second=0, microsecond=0)
        title = "Эта неделя (с " + start_local.strftime("%d.%m") + ")"
    elif period == "month":
        start_local = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        title = "Этот месяц (" + start_local.strftime("%B %Y") + ")"
    else:
        raise ValueError(period)

    # Конвертируем границы в UTC и в тот же ISO-формат что хранится в ts
    start_utc = start_local.astimezone(timezone.utc).replace(tzinfo=None)
    end_utc = now_local.astimezone(timezone.utc).replace(tzinfo=None)
    return start_utc.isoformat(), end_utc.isoformat(), title


def get_period_report(chat_id: int, period: str) -> dict:
    """
    Сводка по всем операциям за период (day/week/month) в локальном времени.
    """
    start_iso, end_iso, title = _period_bounds(period)
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
            "AND ts >= ? AND ts <= ? ORDER BY id",
            (chat_id, start_iso, end_iso)).fetchall()

    agg = {
        "title": title, "period": period, "count": len(rows),
        "sell_usdt": 0.0, "sell_rub": 0.0, "sell_count": 0,
        "buy_usdt": 0.0, "buy_rub": 0.0, "buy_count": 0,
        "arb_profit": 0.0, "arb_count": 0,
        "cash_in": 0.0, "cash_out": 0.0,
        "loan_out": 0.0, "loan_in": 0.0, "repay_in": 0.0, "repay_out": 0.0,
        "clients": set(),
    }
    for r in rows:
        t = r["type"]
        cp = r["counterparty"]
        if t == "sell":
            agg["sell_usdt"] += r["usdt"] or 0; agg["sell_rub"] += r["rubles"] or 0
            agg["sell_count"] += 1
            if cp: agg["clients"].add(name_root(cp))
        elif t == "buy":
            agg["buy_usdt"] += r["usdt"] or 0; agg["buy_rub"] += r["rubles"] or 0
            agg["buy_count"] += 1
            if cp: agg["clients"].add(name_root(cp))
        elif t == "arb":
            agg["arb_profit"] += r["usdt"] or 0; agg["arb_count"] += 1
        elif t == "cash_in":
            agg["cash_in"] += r["rubles"] or 0
        elif t == "cash_out":
            agg["cash_out"] += r["rubles"] or 0
        elif t == "loan_out":
            agg["loan_out"] += r["rubles"] or 0
        elif t == "loan_in":
            agg["loan_in"] += r["rubles"] or 0
        elif t == "debt_repay_in":
            agg["repay_in"] += r["rubles"] or 0
        elif t == "debt_repay_out":
            agg["repay_out"] += r["rubles"] or 0

    agg["avg_sell"] = (agg["sell_rub"] / agg["sell_usdt"]) if agg["sell_usdt"] else None
    agg["avg_buy"] = (agg["buy_rub"] / agg["buy_usdt"]) if agg["buy_usdt"] else None
    agg["spread"] = (agg["avg_sell"] - agg["avg_buy"]) if (agg["avg_sell"] and agg["avg_buy"]) else None
    # Реализованная прибыль за период (по сделкам периода)
    if agg["avg_sell"] and agg["avg_buy"]:
        traded = min(agg["sell_usdt"], agg["buy_usdt"])
        agg["realized_profit"] = traded * (agg["avg_sell"] - agg["avg_buy"])
    else:
        agg["realized_profit"] = None
    agg["n_clients"] = len(agg["clients"])
    return agg


def format_period_report(agg: dict) -> str:
    """Текст отчёта за период."""
    lines = [f"📅 *Отчёт — {agg['title']}*"]
    if agg["count"] == 0:
        lines.append("\nОпераций не было.")
        return "\n".join(lines)
    lines.append(f"_{agg['count']} операций_\n")

    # USDT-сделки
    if agg["sell_count"] or agg["buy_count"]:
        lines.append("📊 *USDT-сделки:*")
        if agg["sell_count"]:
            lines.append(f"   📤 Продано: {fmt_usdt(agg['sell_usdt'])} "
                         f"на {fmt_rub(agg['sell_rub'])}"
                         + (f" (ср. {fmt_rate(agg['avg_sell'])})" if agg["avg_sell"] else ""))
        if agg["buy_count"]:
            lines.append(f"   📥 Куплено: {fmt_usdt(agg['buy_usdt'])} "
                         f"за {fmt_rub(agg['buy_rub'])}"
                         + (f" (ср. {fmt_rate(agg['avg_buy'])})" if agg["avg_buy"] else ""))
        if agg["spread"] is not None:
            sign = "+" if agg["spread"] >= 0 else ""
            lines.append(f"   📈 Спред: {sign}{agg['spread']:.3f}₽/USDT")
        if agg["realized_profit"] is not None:
            ps = "+" if agg["realized_profit"] >= 0 else "−"
            lines.append(f"   💵 Прибыль с оборота: {ps}{fmt_rub(abs(agg['realized_profit']))}")
        if agg["n_clients"]:
            lines.append(f"   👥 Клиентов: {agg['n_clients']}")

    if agg["arb_count"]:
        lines.append(f"\n🔄 *Арбитраж:* {agg['arb_count']} сделок, "
                     f"профит +{fmt_usdt(agg['arb_profit'])}")

    # Кэш
    if agg["cash_in"] or agg["cash_out"]:
        lines.append("\n💵 *Кэш:*")
        if agg["cash_in"]:
            lines.append(f"   Приход: +{fmt_rub(agg['cash_in'])}")
        if agg["cash_out"]:
            lines.append(f"   Расход: −{fmt_rub(agg['cash_out'])}")
        net = agg["cash_in"] - agg["cash_out"]
        ns = "+" if net >= 0 else "−"
        lines.append(f"   Итого: {ns}{fmt_rub(abs(net))}")

    # Долги
    if agg["loan_out"] or agg["loan_in"] or agg["repay_in"] or agg["repay_out"]:
        lines.append("\n💳 *Долги за период:*")
        if agg["loan_out"]:
            lines.append(f"   Выдал в долг: {fmt_rub(agg['loan_out'])}")
        if agg["repay_in"]:
            lines.append(f"   Вернули нам: {fmt_rub(agg['repay_in'])}")
        if agg["loan_in"]:
            lines.append(f"   Заняли мы: {fmt_rub(agg['loan_in'])}")
        if agg["repay_out"]:
            lines.append(f"   Вернули мы: {fmt_rub(agg['repay_out'])}")

    return "\n".join(lines)


# ───── Экспорт в Excel ─────
TYPE_LABELS = {
    "sell": "Продажа USDT",
    "buy": "Покупка USDT",
    "cash_in": "Приход кэш",
    "cash_out": "Расход кэш",
    "arb": "Арбитраж",
    "loan_out": "Дал в долг",
    "loan_in": "Занял",
    "debt_repay_in": "Вернули нам долг",
    "debt_repay_out": "Вернули мы долг",
}


def _ts_to_local_str(ts: str) -> str:
    """ISO-строку UTC → локальное время пользователя 'ДД.ММ.ГГГГ ЧЧ:ММ'."""
    if not ts:
        return ""
    try:
        dt = datetime.fromisoformat(ts)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(LOCAL_TZ).strftime("%d.%m.%Y %H:%M")
    except (ValueError, TypeError):
        return ts


def build_excel_export(chat_id: int, period: str = None) -> str:
    """
    Строит .xlsx со всеми операциями (+ клиенты, долги, сводка) и возвращает путь.
    period: None = всё, иначе 'day'/'week'/'month' — фильтр операций по периоду.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Выбираем операции
    if period:
        start_iso, end_iso, ptitle = _period_bounds(period)
        with db() as conn:
            rows = conn.execute(
                "SELECT * FROM transactions WHERE chat_id = ? AND ts >= ? AND ts <= ? "
                "ORDER BY id", (chat_id, start_iso, end_iso)).fetchall()
        title_suffix = ptitle
    else:
        with db() as conn:
            rows = conn.execute(
                "SELECT * FROM transactions WHERE chat_id = ? ORDER BY id",
                (chat_id,)).fetchall()
        title_suffix = "Все операции"

    wb = Workbook()
    FONT = "Arial"
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2F5496")
    title_font = Font(name=FONT, bold=True, size=14)
    normal_font = Font(name=FONT, size=10)
    bold_font = Font(name=FONT, bold=True, size=10)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '#,##0 ₽'
    usdt_fmt = '#,##0.00'
    rate_fmt = '0.0000'

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    # ── Лист 1: Операции ──
    ws = wb.active
    ws.title = "Операции"
    ws["A1"] = f"TUZ УЧЁТ — {title_suffix}"
    ws["A1"].font = title_font
    ws["A2"] = f"Сформировано: {datetime.now(LOCAL_TZ).strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

    headers = ["ID", "Дата/время", "Тип", "Контрагент", "USDT", "Курс", "Сумма ₽", "Статус", "Исходный текст"]
    hrow = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, len(headers), hrow)

    r = hrow + 1
    for row in rows:
        ws.cell(row=r, column=1, value=f"#{row['id']:03d}").font = normal_font
        ws.cell(row=r, column=2, value=_ts_to_local_str(row["ts"])).font = normal_font
        ws.cell(row=r, column=3, value=TYPE_LABELS.get(row["type"], row["type"])).font = normal_font
        ws.cell(row=r, column=4, value=row["counterparty"] or "—").font = normal_font
        c5 = ws.cell(row=r, column=5, value=round(row["usdt"], 2) if row["usdt"] else None)
        c5.font = normal_font; c5.number_format = usdt_fmt
        c6 = ws.cell(row=r, column=6, value=round(row["rate"], 4) if row["rate"] else None)
        c6.font = normal_font; c6.number_format = rate_fmt
        c7 = ws.cell(row=r, column=7, value=round(row["rubles"], 2) if row["rubles"] else None)
        c7.font = normal_font; c7.number_format = money_fmt
        st = "✓" if (row["status"] if "status" in row.keys() else "confirmed") == "confirmed" else "⏳ ждёт"
        ws.cell(row=r, column=8, value=st).font = normal_font
        ws.cell(row=r, column=9, value=row["raw_text"] or "").font = Font(name=FONT, size=9, color="808080")
        r += 1

    widths = [8, 18, 18, 22, 12, 10, 16, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A5"

    # ── Лист 2: Клиенты ──
    clients = get_all_clients(chat_id)
    if clients:
        wc = wb.create_sheet("Клиенты")
        wc["A1"] = "Клиенты по обороту USDT-сделок"
        wc["A1"].font = title_font
        chead = ["Клиент", "Сделок", "Продаж", "Покупок", "Оборот ₽", "Оборот USDT", "Ср. продажа", "Ср. закупка", "Спред ₽"]
        for i, h in enumerate(chead, 1):
            wc.cell(row=3, column=i, value=h)
        style_header(wc, len(chead), 3)
        cr = 4
        for c in clients:
            wc.cell(row=cr, column=1, value=c["display_name"]).font = bold_font
            wc.cell(row=cr, column=2, value=c["ops"]).font = normal_font
            wc.cell(row=cr, column=3, value=c["sell_count"]).font = normal_font
            wc.cell(row=cr, column=4, value=c["buy_count"]).font = normal_font
            x5 = wc.cell(row=cr, column=5, value=round(c["turnover_rub"], 2)); x5.font = normal_font; x5.number_format = money_fmt
            x6 = wc.cell(row=cr, column=6, value=round(c["turnover_usdt"], 2)); x6.font = normal_font; x6.number_format = usdt_fmt
            x7 = wc.cell(row=cr, column=7, value=round(c["avg_sell"], 4) if c["avg_sell"] else None); x7.font = normal_font; x7.number_format = rate_fmt
            x8 = wc.cell(row=cr, column=8, value=round(c["avg_buy"], 4) if c["avg_buy"] else None); x8.font = normal_font; x8.number_format = rate_fmt
            x9 = wc.cell(row=cr, column=9, value=round(c["spread"], 4) if c["spread"] is not None else None); x9.font = normal_font; x9.number_format = rate_fmt
            cr += 1
        for i, w in enumerate([22, 9, 9, 9, 16, 14, 12, 12, 10], 1):
            wc.column_dimensions[chr(64 + i)].width = w
        wc.freeze_panes = "A4"

    # ── Лист 3: Долги ──
    debts = get_debts(chat_id)
    if debts["owed_to_us"] or debts["we_owe"]:
        wd = wb.create_sheet("Долги")
        wd["A1"] = "Открытые долги"
        wd["A1"].font = title_font
        wd.cell(row=3, column=1, value="Нам должны").font = bold_font
        wd.cell(row=3, column=2, value="Сумма ₽").font = bold_font
        dr = 4
        for name, amt in debts["owed_to_us"]:
            wd.cell(row=dr, column=1, value=name).font = normal_font
            x = wd.cell(row=dr, column=2, value=round(amt, 2)); x.font = normal_font; x.number_format = money_fmt
            dr += 1
        tot = wd.cell(row=dr, column=1, value="ИТОГО нам должны"); tot.font = bold_font
        x = wd.cell(row=dr, column=2, value=round(debts["total_owed_to_us"], 2)); x.font = bold_font; x.number_format = money_fmt
        dr += 2
        wd.cell(row=dr, column=1, value="Мы должны").font = bold_font
        wd.cell(row=dr, column=2, value="Сумма ₽").font = bold_font
        dr += 1
        for name, amt in debts["we_owe"]:
            wd.cell(row=dr, column=1, value=name).font = normal_font
            x = wd.cell(row=dr, column=2, value=round(amt, 2)); x.font = normal_font; x.number_format = money_fmt
            dr += 1
        tot = wd.cell(row=dr, column=1, value="ИТОГО мы должны"); tot.font = bold_font
        x = wd.cell(row=dr, column=2, value=round(debts["total_we_owe"], 2)); x.font = bold_font; x.number_format = money_fmt
        wd.column_dimensions["A"].width = 28
        wd.column_dimensions["B"].width = 18

    # ── Лист 4: Сводка ──
    state = get_state(chat_id)
    rates = get_avg_rates(chat_id, period="all")
    wsum = wb.create_sheet("Сводка")
    wsum["A1"] = "Сводка"
    wsum["A1"].font = title_font
    summary = [
        ("Касса (₽)", round(state["ruble_balance"], 2), money_fmt),
        ("Кошелёк (USDT)", round(state["usdt_balance"], 2), usdt_fmt),
        ("Средняя продажа (всё время)", round(rates["avg_sell"], 4) if rates["avg_sell"] else None, rate_fmt),
        ("Средняя закупка (всё время)", round(rates["avg_buy"], 4) if rates["avg_buy"] else None, rate_fmt),
        ("Спред (₽/USDT)", round(rates["spread"], 4) if rates["spread"] is not None else None, rate_fmt),
        ("Реализованная прибыль (₽)", round(rates["realized_profit_rub"], 2) if rates["realized_profit_rub"] is not None else None, money_fmt),
        ("Профит арбитража (USDT)", round(rates["arb_profit_usdt"], 2) if rates["arb_profit_usdt"] else None, usdt_fmt),
        ("Всего операций в выгрузке", len(rows), "0"),
    ]
    sr = 3
    for label, val, fmt in summary:
        wsum.cell(row=sr, column=1, value=label).font = bold_font
        cell = wsum.cell(row=sr, column=2, value=val)
        cell.font = normal_font
        if fmt:
            cell.number_format = fmt
        sr += 1
    wsum.column_dimensions["A"].width = 32
    wsum.column_dimensions["B"].width = 20

    # Сохраняем во временный файл
    suffix = period or "all"
    out_path = DATA_DIR / f"export_{chat_id}_{suffix}_{datetime.now(LOCAL_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(out_path)
    return str(out_path)


# ────────────────── ПАРСИНГ ──────────────────
NUM_PATTERN = r"\d[\d\s\u00a0.,]*\d|\d"

SELL_VERBS = ("продал", "продала", "продали", "продаю", "продают")
BUY_VERBS = ("купил", "купила", "купили", "куплю", "покупаю",
             "откупил", "откупила", "откупили")

CASH_OUT_VERBS = ("выдал", "выдала", "выдали", "выдавал", "выдавали",
                  "отдал", "отдала", "отдали", "отдавал", "отдавали",
                  "перевёл", "перевел", "перевела", "перевели",
                  "потратил", "потратила", "потратили")
CASH_IN_VERBS = ("принял", "приняла", "приняли", "принимал", "принимали",
                 "забрал", "забрала", "забрали", "забирал", "забирали",
                 "получил", "получила", "получили", "получал", "получали")

ARB_MARKERS = (
    "сделка без моих средств",
    "сделка без своих средств",
    "арбитраж",
    "партнёрская сделка",
    "партнерская сделка",
)


def parse_number(s: str, prefer_decimal: bool = False) -> float:
    """
    Парсит '72 600' → 72600, '85.000' → 85000, '73.5' → 73.5,
    '1 142 400' → 1142400, '75,6' → 75.6, '1.000.000' → 1000000.

    prefer_decimal=True — для курсов: единственный разделитель всегда десятичный.
    """
    s = s.strip().replace(" ", "").replace("\u00a0", "")
    s = s.rstrip("₽рp$ ").strip()
    if not s:
        raise ValueError("empty")

    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif prefer_decimal:
        s = s.replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if (all(len(p) == 3 and p.isdigit() for p in parts[1:])
                and 1 <= len(parts[0]) <= 3 and parts[0].isdigit()):
            s = s.replace(".", "")
    elif "," in s:
        parts = s.split(",")
        if (all(len(p) == 3 and p.isdigit() for p in parts[1:])
                and 1 <= len(parts[0]) <= 3 and parts[0].isdigit()):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")

    return float(s)


def _find_verb(lower: str, verb_groups: list) -> tuple:
    """Ищет глагол. Возвращает (kind, start, end) или (None, None, None)."""
    best = None
    for kind, verbs in verb_groups:
        for v in verbs:
            i = lower.find(v)
            if i < 0 or i > 30:
                continue
            before = lower[:i].strip()
            if before and " " in before:
                continue
            end = i + len(v)
            before_ok = (i == 0) or not lower[i - 1].isalpha()
            after_ok = (end == len(lower)) or not lower[end].isalpha()
            if before_ok and after_ok:
                if best is None or i < best[1]:
                    best = (kind, i, end)
    return best if best else (None, None, None)


def parse_trade(text: str):
    """USDT-сделка. Возвращает dict или None."""
    text = text.strip().lstrip("+-—– ").strip()
    lower = text.lower()

    kind, vs, ve = _find_verb(lower, [("sell", SELL_VERBS), ("buy", BUY_VERBS)])
    if not kind:
        return None

    doer = text[:vs].strip() or None
    rest = text[ve:].lstrip()
    for prep in ("у ", "у\u00a0", "от ", "от\u00a0"):
        if rest.lower().startswith(prep):
            rest = rest[len(prep):].lstrip()
            break

    m = re.search(rf"({NUM_PATTERN})\s*([*x×/÷:])\s*({NUM_PATTERN})", rest)
    if not m:
        return None

    counterparty = rest[:m.start()].strip().rstrip("—-:,.").strip() or "—"
    try:
        op = m.group(2)
        a = parse_number(m.group(1))
        b = parse_number(m.group(3), prefer_decimal=True)
        if op in "*x×":
            usdt, rate, rubles = a, b, a * b
        else:
            rubles, rate, usdt = a, b, a / b
    except (ValueError, ZeroDivisionError):
        return None

    if rate <= 0 or usdt <= 0 or rubles <= 0:
        return None

    return {
        "type": kind,
        "doer": doer,
        "counterparty": counterparty,
        "usdt": round(usdt, 4),
        "rate": rate,
        "rubles": round(rubles, 2),
    }


def parse_cash_flow(text: str):
    """Кэш-движение (только рубли). Возвращает dict или None."""
    text = text.strip()

    explicit_dir = None
    m_sign = re.match(r"^([+\-—–])\s*", text)
    if m_sign:
        explicit_dir = "cash_out" if m_sign.group(1) in "-—–" else "cash_in"
        text = text[m_sign.end():].strip()

    lower = text.lower()

    kind, vs, ve = _find_verb(lower, [
        ("cash_in", CASH_IN_VERBS),
        ("cash_out", CASH_OUT_VERBS),
    ])

    if not kind:
        if not explicit_dir:
            return None
        m_amt = re.search(NUM_PATTERN, text)
        if not m_amt:
            return None
        try:
            amount = parse_number(m_amt.group())
        except ValueError:
            return None
        if amount <= 0:
            return None
        before = text[:m_amt.start()].strip(" ,:.")
        after = text[m_amt.end():].strip(" ,:.")
        for prep in ("от ", "у ", "из ", "к ", "для "):
            if after.lower().startswith(prep):
                after = after[len(prep):].strip()
                break
        counterparty = (before + " " + after).strip().strip(",:.") or "—"
        return {"type": explicit_dir, "doer": None, "counterparty": counterparty, "amount": amount}

    direction = kind
    doer = text[:vs].strip() or None
    rest = text[ve:].lstrip()
    for prep in ("у ", "у\u00a0", "от ", "от\u00a0", "из ", "к ", "для "):
        if rest.lower().startswith(prep):
            rest = rest[len(prep):].lstrip()
            break

    matches = list(re.finditer(NUM_PATTERN, rest))
    if not matches:
        return None
    last_m = matches[-1]
    try:
        amount = parse_number(last_m.group())
    except ValueError:
        return None
    if amount <= 0:
        return None

    target = rest[:last_m.start()].strip().rstrip(",:.").rstrip().rstrip("—-").strip() or "—"
    return {"type": direction, "doer": doer, "counterparty": target, "amount": amount}


def parse_loan(text: str):
    """
    Долговые операции. Возвращает dict или None.

    Шаблоны:
      Дал в долг Михаилу 90000        → loan_out  (касса -, должник Михаил +)
      Занял у Васи 200000             → loan_in   (касса +, кредитор Вася +)
      Михаил вернул 50000             → debt_repay_in  (касса +, долг Михаила -)
      Вернул Васе 100000              → debt_repay_out (касса -, наш долг Васе -)
    """
    t = text.strip().lstrip("+-—– ").strip()
    lower = t.lower()

    # Дал в долг X сумма
    m = re.match(
        r"^(?:дал|дала|дали|выдал|выдала|выдали|занял\s+(?!у\b))\s+в\s+долг\s+(.+?)\s+(" + NUM_PATTERN + r")\s*₽?\s*$",
        t, re.IGNORECASE)
    if m:
        try:
            amt = parse_number(m.group(2))
        except ValueError:
            return None
        if amt <= 0:
            return None
        return {"type": "loan_out", "counterparty": m.group(1).strip(),
                "amount": amt, "doer": None}

    # Занял у X сумма (без "в долг")
    m = re.match(
        r"^(?:занял|заняла|заняли|взял\s+в\s+долг|взяли\s+в\s+долг)\s+(?:у\s+)?(.+?)\s+(" + NUM_PATTERN + r")\s*₽?\s*$",
        t, re.IGNORECASE)
    if m:
        try:
            amt = parse_number(m.group(2))
        except ValueError:
            return None
        if amt <= 0:
            return None
        return {"type": "loan_in", "counterparty": m.group(1).strip(),
                "amount": amt, "doer": None}

    # X вернул [нам] сумма
    m = re.match(
        r"^(.+?)\s+(?:вернул|вернула|вернули|отдал|отдала|отдали|погасил|погасила|погасили)"
        r"(?:\s+(?:нам|мне|долг))?\s+(" + NUM_PATTERN + r")\s*₽?\s*$",
        t, re.IGNORECASE)
    if m:
        try:
            amt = parse_number(m.group(2))
        except ValueError:
            return None
        if amt <= 0:
            return None
        return {"type": "debt_repay_in", "counterparty": m.group(1).strip(),
                "amount": amt, "doer": None}

    # Вернул/Отдал X сумма (мы отдаём свой долг)
    m = re.match(
        r"^(?:вернул|вернула|вернули|отдал\s+долг|отдала\s+долг|отдали\s+долг|погасил|погасила|погасили)"
        r"\s+(.+?)\s+(" + NUM_PATTERN + r")\s*₽?\s*$",
        t, re.IGNORECASE)
    if m:
        try:
            amt = parse_number(m.group(2))
        except ValueError:
            return None
        if amt <= 0:
            return None
        return {"type": "debt_repay_out", "counterparty": m.group(1).strip(),
                "amount": amt, "doer": None}

    return None


def parse_message_line(line: str):
    """Парсит одну строку: сначала сделку, потом долг, потом кэш-движение."""
    line = line.strip()
    line = re.sub(r"^\d+\s*[\.\)]\s*", "", line).strip()
    if not line:
        return None
    t = parse_trade(line)
    if t:
        return t
    l = parse_loan(line)
    if l:
        return l
    c = parse_cash_flow(line)
    if c:
        return c
    return None


def parse_arbitrage_message(text: str):
    """
    Арбитраж: маркер + 2 trade-формат блока.
    Касса не меняется, в кошелёк падает разница USDT (профит).

    Пример:
        Сделка без моих средств

        Продал Стефу
        196500/75.5=2 602,649

        Купил у Германа
        196500/75=2 620

    Возвращает dict или None.
    """
    lower_full = text.lower()
    if not any(m in lower_full for m in ARB_MARKERS):
        return None

    # Делим сообщение на блоки по пустым строкам
    blocks = re.split(r"\n\s*\n", text)

    # В каждом блоке склеиваем строки в одну, пропуская маркер
    trades = []
    for block in blocks:
        lines = []
        for line in block.split("\n"):
            line = re.sub(r"^\d+\s*[\.\)]\s*", "", line.strip()).strip()
            if not line:
                continue
            if any(m in line.lower() for m in ARB_MARKERS):
                continue
            lines.append(line)
        if not lines:
            continue
        joined = " ".join(lines)
        t = parse_trade(joined)
        if t:
            trades.append((joined, t))

    # Если блочное деление дало меньше 2 сделок (например, обе сделки были
    # в одном блоке без пустой строки между ними) — пробуем построчно.
    if len(trades) < 2:
        line_trades = []
        for line in text.split("\n"):
            line = re.sub(r"^\d+\s*[\.\)]\s*", "", line.strip()).strip()
            if not line or any(m in line.lower() for m in ARB_MARKERS):
                continue
            t = parse_trade(line)
            if t:
                line_trades.append((line, t))
        if len(line_trades) >= 2:
            trades = line_trades

    if len(trades) < 2:
        return None

    # Берём первые 2 сделки
    (raw_a, a), (raw_b, b) = trades[0], trades[1]

    # Та, что больше USDT — это что мы получили; меньшая — что отдали.
    if a["usdt"] >= b["usdt"]:
        got, sent = a, b
    else:
        got, sent = b, a

    profit = round(got["usdt"] - sent["usdt"], 4)

    # Sanity check: профит не должен быть нулевым/отрицательным и не больше 30% от объёма
    if profit <= 0 or profit > got["usdt"] * 0.30:
        return None

    return {
        "type": "arb",
        "partner_in": got["counterparty"],
        "partner_out": sent["counterparty"],
        "usdt_in": got["usdt"],
        "usdt_out": sent["usdt"],
        "rub_amount": round(max(got["rubles"], sent["rubles"]), 2),
        "rate_in": got["rate"],
        "rate_out": sent["rate"],
        "profit": profit,
    }


# ────────────────── ФОРМАТИРОВАНИЕ ──────────────────
def fmt_rub(n: float) -> str:
    sign = "-" if n < 0 else ""
    n = abs(n)
    s = f"{n:,.2f}".replace(",", " ")
    if s.endswith(".00"):
        s = s[:-3]
    return f"{sign}{s}₽"


def fmt_usdt(n: float) -> str:
    sign = "-" if n < 0 else ""
    n = abs(n)
    s = f"{n:,.2f}".replace(",", " ")
    if s.endswith(".00"):
        s = s[:-3]
    return f"{sign}{s} USDT"


def fmt_rate(n: float) -> str:
    s = f"{n:.4f}".rstrip("0").rstrip(".")
    return s or "0"


def fmt_id(i: int) -> str:
    return f"#{i:03d}"


def format_debt_status(d: dict, focus_name: str = None, max_items: int = 4) -> str:
    """
    Возвращает строку-напоминание о долгах для добавления к ответу бота.
    focus_name — если указан, выводит ТОЛЬКО детали по этому контрагенту.
    Возвращает '' если долгов нет.
    """
    lines = []
    if focus_name:
        focus_lower = focus_name.lower()
        # Ищем должника
        for name, amt in d.get("owed_to_us", []):
            if name.lower() == focus_lower or focus_lower in name.lower():
                lines.append(f"   📥 *{name}* теперь должен нам *{fmt_rub(amt)}*")
                break
        for name, amt in d.get("we_owe", []):
            if name.lower() == focus_lower or focus_lower in name.lower():
                lines.append(f"   📤 *Мы должны {name}: {fmt_rub(amt)}*")
                break
        return "\n".join(lines) if lines else ""

    # Общий вид (для напоминания после сделки)
    owed = d.get("owed_to_us", [])
    we_owe = d.get("we_owe", [])
    if not owed and not we_owe:
        return ""

    if owed:
        names = ", ".join(f"{n} {fmt_rub(a)}" for n, a in owed[:max_items])
        if len(owed) > max_items:
            names += f" + ещё {len(owed) - max_items}"
        lines.append(f"⚠️ Нам должны *{fmt_rub(d['total_owed_to_us'])}*: {names}")
    if we_owe:
        names = ", ".join(f"{n} {fmt_rub(a)}" for n, a in we_owe[:max_items])
        if len(we_owe) > max_items:
            names += f" + ещё {len(we_owe) - max_items}"
        lines.append(f"⚠️ Мы должны *{fmt_rub(d['total_we_owe'])}*: {names}")
    return "\n" + "\n".join(lines)


def fmt_tx_short(r) -> str:
    tid = fmt_id(r["id"])
    cp = r["counterparty"] or "—"
    t = r["type"]
    pending = " ⏳" if (r["status"] if "status" in r.keys() else "confirmed") == "pending" else ""
    if t == "sell":
        return f"{tid}{pending} 📤 Продал {cp}: {fmt_usdt(r['usdt'])} × {fmt_rate(r['rate'])} = {fmt_rub(r['rubles'])}"
    if t == "buy":
        return f"{tid}{pending} 📥 Купил у {cp}: {fmt_usdt(r['usdt'])} × {fmt_rate(r['rate'])} = {fmt_rub(r['rubles'])}"
    if t == "cash_out":
        return f"{tid}{pending} 💸 Выдал {cp}: −{fmt_rub(r['rubles'])}"
    if t == "cash_in":
        return f"{tid}{pending} 💰 Принял {cp}: +{fmt_rub(r['rubles'])}"
    if t == "arb":
        pin = r["partner_in"] if "partner_in" in r.keys() else "?"
        pout = r["partner_out"] if "partner_out" in r.keys() else "?"
        return f"{tid}{pending} 🔄 Арбитраж {pin}→{pout}: +{fmt_usdt(r['usdt'])} профит"
    if t == "loan_out":
        return f"{tid}{pending} 📤💳 Дал в долг {cp}: −{fmt_rub(r['rubles'])}"
    if t == "loan_in":
        return f"{tid}{pending} 📥💳 Занял у {cp}: +{fmt_rub(r['rubles'])}"
    if t == "debt_repay_in":
        return f"{tid}{pending} ↩️💰 {cp} вернул долг: +{fmt_rub(r['rubles'])}"
    if t == "debt_repay_out":
        return f"{tid}{pending} ↩️💸 Вернул {cp}: −{fmt_rub(r['rubles'])}"
    return f"{tid}{pending} {t}: {fmt_rub(r['rubles'])}"


# ────────────────── ЛОГИКА ПРИМЕНЕНИЯ БАЛАНСА ──────────────────
def apply_balance(chat_id: int, tx_row) -> None:
    """Применяет влияние подтверждённой транзакции на кассу/кошелёк."""
    t = tx_row["type"]
    if t == "sell":
        update_state(chat_id, d_rubles=tx_row["rubles"], d_usdt=-tx_row["usdt"])
    elif t == "buy":
        update_state(chat_id, d_rubles=-tx_row["rubles"], d_usdt=tx_row["usdt"])
    elif t == "cash_in":
        update_state(chat_id, d_rubles=tx_row["rubles"])
    elif t == "cash_out":
        update_state(chat_id, d_rubles=-tx_row["rubles"])
    elif t == "arb":
        # Арбитраж НЕ трогает кассу и кошелёк USDT.
        # Профит копится в отдельном счётчике "доход с арбитража".
        update_state(chat_id, d_arb_profit=tx_row["usdt"])
    elif t == "loan_out":          # дали в долг → касса -
        update_state(chat_id, d_rubles=-tx_row["rubles"])
    elif t == "loan_in":           # заняли → касса +
        update_state(chat_id, d_rubles=tx_row["rubles"])
    elif t == "debt_repay_in":     # нам вернули → касса +
        update_state(chat_id, d_rubles=tx_row["rubles"])
    elif t == "debt_repay_out":    # мы вернули → касса -
        update_state(chat_id, d_rubles=-tx_row["rubles"])


def reverse_balance(chat_id: int, tx_row) -> None:
    """Откатывает влияние транзакции."""
    t = tx_row["type"]
    if t == "sell":
        update_state(chat_id, d_rubles=-tx_row["rubles"], d_usdt=tx_row["usdt"])
    elif t == "buy":
        update_state(chat_id, d_rubles=tx_row["rubles"], d_usdt=-tx_row["usdt"])
    elif t == "cash_in":
        update_state(chat_id, d_rubles=-tx_row["rubles"])
    elif t == "cash_out":
        update_state(chat_id, d_rubles=tx_row["rubles"])
    elif t == "arb":
        update_state(chat_id, d_arb_profit=-tx_row["usdt"])
    elif t == "loan_out":
        update_state(chat_id, d_rubles=tx_row["rubles"])
    elif t == "loan_in":
        update_state(chat_id, d_rubles=-tx_row["rubles"])
    elif t == "debt_repay_in":
        update_state(chat_id, d_rubles=-tx_row["rubles"])
    elif t == "debt_repay_out":
        update_state(chat_id, d_rubles=tx_row["rubles"])


def should_confirm(parsed: dict, settings: dict) -> bool:
    """Нужно ли подтверждать через кнопки?"""
    mode = settings.get("confirm_mode", "trades")
    if mode == "off":
        return False
    if mode == "all":
        return True
    t = parsed["type"]
    if mode == "trades":
        # Подтверждаем только сделки и арбитраж. Кэш и долги — сразу.
        return t in ("sell", "buy", "arb")
    if mode == "big":
        amt = parsed.get("rubles") or parsed.get("amount") or parsed.get("rub_amount") or 0
        return amt >= 100000
    return False


# ────────────────── БОТ ──────────────────
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


HELP_MAIN = (
    "👋 Я веду учёт USDT-сделок, кэш-движений, долгов и арбитража.\n"
    "_Тип чата:_ *main* (твой личный — пишешь все операции)\n\n"
    "📝 *USDT-сделки* (меняют ₽ и USDT, требуют подтверждения):\n"
    "`Продал Гиге 475600/76.262`\n"
    "`Купил у Стефа 1020*74`\n"
    "`Влад продал Клиенту 96000/75.7`\n"
    "Правило: `*` слева USDT, `/` слева рубли.\n\n"
    "💵 *Кэш-движения* (меняют только ₽):\n"
    "`Выдал Владу 72 600`\n"
    "`Принял от Адахана 729 100`\n"
    "`+ 96000 от клиента`\n\n"
    "💳 *Долги:*\n"
    "`Дал в долг Михаилу 90000`\n"
    "`Занял у Васи 200000`\n"
    "`Михаил вернул 50000`\n"
    "`Вернул Васе 100000`\n\n"
    "🔄 *Арбитраж* (касса не меняется, в кошелёк падает спред):\n"
    "```\nСделка без моих средств\n"
    "Купил у Германа 196500/75=2620\n"
    "Продал Стефу 196500/75.5=2602\n```\n"
    "👤 *Клиенты:* пиши имя или @username в сделке —\n"
    "`Продал @vasya_crypto 1000*76` — бот сам ведёт по клиенту статистику.\n\n"
    "📊 *Команды:*\n"
    "/balance · /stats · /rates · /history · /cashflow\n"
    "/debts · /debt · /client · /clients · /ct\n"
    "/day · /week · /month _(отчёты за период)_\n"
    "/export · /backup _(выгрузка Excel и копия базы)_\n"
    "/reminders · /setpaydate _(автонапоминания)_\n"
    "/find · /undo · /pending · /setcash · /settype · /confirm · /help"
)

HELP_FIELD = (
    "👋 Это рабочий чат сотрудника (поле).\n"
    "_Тип чата:_ *field* — операции сотрудника\n\n"
    "📝 *Сделки в поле:*\n"
    "`Влад купил у клиента 78200/73.5`\n"
    "`Влад продал клиенту 96000/75.7`\n\n"
    "💵 *Кэш на руках:*\n"
    "`Выдали Владу 72 600` _(из общей кассы → на руки Владу)_\n"
    "`Принял от клиента 96 000` _(клиент дал нал)_\n"
    "`Влад выдал клиенту 30 000`\n\n"
    "💳 *Долги клиентов:*\n"
    "`Дал в долг Серёге 50000`\n"
    "`Серёга вернул 50000`\n\n"
    "📊 *Команды:*\n"
    "/balance · /history · /cashflow · /rates · /debts · /client · /clients · /day · /week · /month · /export · /backup · /find · /undo · /pending · /settype · /help"
)

HELP_COMMON = (
    "👋 Это чат общей кассы.\n"
    "_Тип чата:_ *common* — приходы и расходы общего котла\n\n"
    "💵 *Шаблоны:*\n"
    "`Принял от Влада 500 000` _(Влад сдал в кассу)_\n"
    "`Принял от Ивана 800 000` _(Иван внёс в кассу)_\n"
    "`Выдал Владу 100 000` _(касса → Владу на оборотку)_\n"
    "`Выдал на расходы 30 000`\n\n"
    "💳 *Долги партнёров:*\n"
    "`Дал в долг MTX 1 000 000`\n"
    "`MTX вернул 500 000`\n\n"
    "📊 *Команды:*\n"
    "/balance · /cashflow · /rates · /debts · /client · /clients · /day · /week · /month · /export · /backup · /find · /undo · /history · /settype · /help"
)


def help_text(chat_id: int) -> str:
    s = get_chat_settings(chat_id)
    if s["chat_type"] == "field":
        return HELP_FIELD
    if s["chat_type"] == "common":
        return HELP_COMMON
    return HELP_MAIN


def confirm_kb(batch_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Записать", callback_data=f"c:{batch_id}"),
        InlineKeyboardButton(text="❌ Отмена",   callback_data=f"x:{batch_id}"),
    ]])


@dp.message(CommandStart())
@dp.message(Command("help"))
async def on_start(message: Message):
    await message.answer(help_text(message.chat.id), parse_mode="Markdown")


@dp.message(Command("settype"))
async def on_settype(message: Message):
    parts = message.text.split(maxsplit=1)
    s = get_chat_settings(message.chat.id)
    if len(parts) < 2:
        await message.answer(
            f"Сейчас тип чата: *{s['chat_type']}*\n\n"
            "Доступные:\n"
            "• `/settype main` — твой личный чат (все операции)\n"
            "• `/settype field` — чат сотрудника в поле\n"
            "• `/settype common` — общая касса\n\n"
            "Тип меняет только подсказки и шаблоны в `/help`, а команды и парсер работают одинаково.",
            parse_mode="Markdown")
        return
    new_type = parts[1].strip().lower()
    if new_type not in ("main", "field", "common"):
        await message.answer("Тип должен быть main, field или common.")
        return
    set_chat_setting(message.chat.id, chat_type=new_type)
    await message.answer(f"✅ Тип чата: *{new_type}*\n\nНовая памятка — /help", parse_mode="Markdown")


@dp.message(Command("confirm"))
async def on_confirm_cmd(message: Message):
    parts = message.text.split(maxsplit=1)
    s = get_chat_settings(message.chat.id)
    if len(parts) < 2:
        await message.answer(
            f"Сейчас режим подтверждения: *{s['confirm_mode']}*\n\n"
            "Доступные:\n"
            "• `/confirm trades` — кнопки только для USDT-сделок и арбитража _(по умолчанию)_\n"
            "• `/confirm all` — кнопки для всего, включая кэш\n"
            "• `/confirm big` — кнопки только для сделок от 100 000₽\n"
            "• `/confirm off` — без кнопок, записывать сразу",
            parse_mode="Markdown")
        return
    mode = parts[1].strip().lower()
    if mode not in ("all", "trades", "big", "off"):
        await message.answer("Режим должен быть all, trades, big или off.")
        return
    set_chat_setting(message.chat.id, confirm_mode=mode)
    await message.answer(f"✅ Подтверждение: *{mode}*", parse_mode="Markdown")


@dp.message(Command("balance"))
async def on_balance(message: Message):
    s = get_state(message.chat.id)
    text = f"💰 *Касса:* {fmt_rub(s['ruble_balance'])}\n💵 *Кошелёк:* {fmt_usdt(s['usdt_balance'])}"
    arb_total = s.get("arb_profit_usdt", 0) or 0
    if arb_total:
        text += f"\n💼 *Доход с арбитража:* {fmt_usdt(arb_total)}"
    if s["extra_rubles"]:
        text += f"\n♻️ *Лишние:* {fmt_rub(s['extra_rubles'])}"
    # Сколько pending
    with db() as conn:
        pending = conn.execute(
            "SELECT COUNT(*) AS c FROM transactions WHERE chat_id = ? AND status = 'pending'",
            (message.chat.id,)).fetchone()["c"]
    if pending:
        text += f"\n⏳ Не подтверждено: *{pending}* (см. /pending)"

    # Краткая сводка по курсам текущего цикла
    cycle = get_avg_rates(message.chat.id, period="cycle")
    if cycle["avg_buy"] or cycle["avg_sell"]:
        text += "\n\n📊 *Цикл:*"
        if cycle["avg_buy"]:
            text += f"\n   📥 Ср. закупка: {fmt_rate(cycle['avg_buy'])}"
        if cycle["avg_sell"]:
            text += f"\n   📤 Ср. продажа: {fmt_rate(cycle['avg_sell'])}"
        if cycle["spread"] is not None:
            sign = "+" if cycle["spread"] >= 0 else ""
            text += f"\n   📈 Спред: {sign}{cycle['spread']:.3f}₽/USDT"
        text += "\n   _Подробнее: /rates_"

    await message.answer(text, parse_mode="Markdown")


@dp.message(Command("stats"))
async def on_stats(message: Message):
    r = get_avg_rates(message.chat.id)
    s = get_state(message.chat.id)

    if not r["avg_sell"] and not r["avg_buy"] and not r["arb_profit_usdt"]:
        await message.answer("USDT-сделок ещё нет.")
        return

    lines = ["📊 *Статистика USDT-сделок*\n"]
    if r["avg_sell"]:
        lines.append(f"📈 Средний курс ПРОДАЖИ: *{fmt_rate(r['avg_sell'])}*")
        lines.append(f"   Всего продано: {fmt_usdt(r['total_sold_usdt'])} → {fmt_rub(r['total_received'])}")
    if r["avg_buy"]:
        lines.append(f"📉 Средний курс ПОКУПКИ: *{fmt_rate(r['avg_buy'])}*")
        lines.append(f"   Всего куплено: {fmt_usdt(r['total_bought_usdt'])} за {fmt_rub(r['total_spent'])}")

    if r["avg_sell"] and r["avg_buy"]:
        spread = r["avg_sell"] - r["avg_buy"]
        realized = (r["avg_sell"] - r["avg_buy"]) * min(r["total_sold_usdt"], r["total_bought_usdt"])
        lines.append(f"\n💹 Спред: *{spread:+.3f}* ₽/USDT")
        lines.append(f"💵 Реализованная прибыль ≈ *{fmt_rub(realized)}*")
        lines.append("\n⚠️ Чтобы не уйти в минус:")
        lines.append(f"   • Покупай НЕ дороже *{fmt_rate(r['avg_sell'])}*")
        lines.append(f"   • Продавай НЕ дешевле *{fmt_rate(r['avg_buy'])}*")

    if r["arb_profit_usdt"]:
        lines.append(f"\n🔄 Профит от арбитража: *+{fmt_usdt(r['arb_profit_usdt'])}*")

    lines.append(f"\n💰 Касса: {fmt_rub(s['ruble_balance'])}")
    lines.append(f"💵 Кошелёк: {fmt_usdt(s['usdt_balance'])}")
    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("rates"))
async def on_rates(message: Message):
    """
    /rates           — текущий цикл
    /rates cycle     — текущий цикл (явно)
    /rates all       — всё время
    /rates 30d       — последние N дней (3d, 7d, 30d, 90d ...)
    """
    parts = message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else "cycle"

    period = "cycle"
    days = 30
    period_title = "Текущий цикл"

    if arg == "all":
        period = "all"
        period_title = "Всё время"
    elif arg == "cycle":
        period = "cycle"
        period_title = "Текущий цикл"
    elif re.match(r"^\d+d$", arg):
        period = "days"
        days = int(arg[:-1])
        period_title = f"Последние {days} дн."
    else:
        await message.answer(
            "Использование: `/rates` `[cycle|all|30d]`\n"
            "Примеры:\n"
            "`/rates`        — текущий цикл (по умолчанию)\n"
            "`/rates all`    — всё время\n"
            "`/rates 30d`    — последние 30 дней\n"
            "`/rates 7d`     — последние 7 дней",
            parse_mode="Markdown")
        return

    r = get_avg_rates(message.chat.id, period=period, days=days)
    s = get_state(message.chat.id)

    if not r["avg_sell"] and not r["avg_buy"] and not r["arb_profit_usdt"]:
        await message.answer(f"📊 *{period_title}*\n\nUSDT-сделок в этом периоде нет.",
                             parse_mode="Markdown")
        return

    lines = [f"📊 *Курсы — {period_title}*",
             f"_({r['tx_count']} USDT-операций)_\n"]

    if r["avg_buy"]:
        lines.append(f"📥 *Средняя закупка:* {fmt_rate(r['avg_buy'])}")
        lines.append(f"   {fmt_usdt(r['total_bought_usdt'])} за {fmt_rub(r['total_spent'])}")
    if r["avg_sell"]:
        lines.append(f"📤 *Средняя продажа:* {fmt_rate(r['avg_sell'])}")
        lines.append(f"   {fmt_usdt(r['total_sold_usdt'])} → {fmt_rub(r['total_received'])}")

    if r["spread"] is not None:
        spread_sign = "+" if r["spread"] >= 0 else ""
        spread_emoji = "📈" if r["spread"] >= 0 else "📉"
        lines.append(f"\n{spread_emoji} *Спред: {spread_sign}{r['spread']:.3f} ₽/USDT*")

    if r["realized_profit_rub"] is not None:
        profit_sign = "+" if r["realized_profit_rub"] >= 0 else "−"
        lines.append(f"💵 *Реализованная прибыль: {profit_sign}{fmt_rub(abs(r['realized_profit_rub']))}*")

    if r["arb_profit_usdt"]:
        lines.append(f"🔄 Профит от арбитража: +{fmt_usdt(r['arb_profit_usdt'])}")

    # ⚖️ Безубыток — только если есть и закупки, и продажи
    if r["avg_sell"] and r["avg_buy"]:
        lines.append(f"\n⚖️ *Безубыток:*")
        lines.append(f"   • Продавай НЕ дешевле *{fmt_rate(r['avg_buy'])}*  (≤ средняя закупка = убыток)")
        lines.append(f"   • Покупай НЕ дороже *{fmt_rate(r['avg_sell'])}*   (≥ средняя продажа = убыток)")

    # Оценка текущей позиции
    if s["usdt_balance"] > 0.01 and r["avg_sell"]:
        potential = s["usdt_balance"] * r["avg_sell"]
        lines.append(f"\n💼 *Сейчас в кошельке:* {fmt_usdt(s['usdt_balance'])}")
        lines.append(f"   По средней продаже = ~{fmt_rub(potential)}")

    # Подсказка
    if period == "cycle":
        lines.append("\n_Цикл = с момента когда кошелёк был ≤ 10 USDT_")
        lines.append("_`/rates all` — всё время, `/rates 30d` — за 30 дней_")

    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("history"))
async def on_history(message: Message):
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
            "ORDER BY id DESC LIMIT 15",
            (message.chat.id,)).fetchall()
    if not rows:
        await message.answer("Операций ещё нет.")
        return
    lines = ["📜 *Последние операции:*\n"]
    for r in rows:
        lines.append(fmt_tx_short(r))
    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("cashflow"))
async def on_cashflow(message: Message):
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? "
            "AND type IN ('cash_in', 'cash_out') AND status = 'confirmed' "
            "ORDER BY id DESC LIMIT 20",
            (message.chat.id,)).fetchall()
    if not rows:
        await message.answer("Кэш-движений ещё нет.")
        return
    lines = ["💵 *Последние кэш-движения:*\n"]
    for r in rows:
        sign = "−" if r["type"] == "cash_out" else "+"
        cp = r["counterparty"] or "—"
        lines.append(f"{fmt_id(r['id'])} {sign}{fmt_rub(r['rubles'])} — {cp}")
    s = get_state(message.chat.id)
    lines.append(f"\n💰 Текущая касса: *{fmt_rub(s['ruble_balance'])}*")
    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("debts"))
async def on_debts(message: Message):
    d = get_debts(message.chat.id)
    lines = ["💳 *Долги*\n"]

    if d["owed_to_us"]:
        lines.append(f"📥 *Нам должны:* {fmt_rub(d['total_owed_to_us'])}")
        for name, amt in d["owed_to_us"]:
            lines.append(f"   • {name}: {fmt_rub(amt)}")
    else:
        lines.append("📥 *Нам никто не должен.*")

    lines.append("")

    if d["we_owe"]:
        lines.append(f"📤 *Мы должны:* {fmt_rub(d['total_we_owe'])}")
        for name, amt in d["we_owe"]:
            lines.append(f"   • {name}: {fmt_rub(amt)}")
    else:
        lines.append("📤 *Мы никому не должны.*")

    net = d["total_owed_to_us"] - d["total_we_owe"]
    lines.append("")
    if net > 0:
        lines.append(f"📊 *Чистый баланс: +{fmt_rub(net)}* (нам должны больше, чем мы)")
    elif net < 0:
        lines.append(f"📊 *Чистый баланс: −{fmt_rub(abs(net))}* (мы должны больше, чем нам)")
    else:
        lines.append("📊 *Чистый баланс: 0*")

    lines.append("\nДеталь по человеку: `/debt <имя>`")
    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("debt"))
async def on_debt(message: Message):
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer(
            "Использование: `/debt <имя>`\n"
            "Например: `/debt Михаил`",
            parse_mode="Markdown")
        return
    name = parts[1].strip()
    rows = get_debt_history(message.chat.id, name)
    if not rows:
        await message.answer(f"Долговых операций по «{name}» не нашёл.")
        return

    # Считаем остаток: для loan_out / debt_repay_in — нам должны
    # для loan_in / debt_repay_out — мы должны
    bal_owed = 0.0  # нам должны
    bal_owe = 0.0   # мы должны
    for r in rows:
        if r["type"] == "loan_out":
            bal_owed += r["rubles"]
        elif r["type"] == "debt_repay_in":
            bal_owed -= r["rubles"]
        elif r["type"] == "loan_in":
            bal_owe += r["rubles"]
        elif r["type"] == "debt_repay_out":
            bal_owe -= r["rubles"]

    lines = [f"💳 *История долгов по «{name}»* ({len(rows)} оп.)\n"]
    for r in rows:
        lines.append(fmt_tx_short(r))
    lines.append("")
    if abs(bal_owed) > 0.01:
        if bal_owed > 0:
            lines.append(f"📥 *Остаток: должен нам {fmt_rub(bal_owed)}*")
        else:
            lines.append(f"⚠️ Переплата: нам вернули на {fmt_rub(-bal_owed)} больше")
    if abs(bal_owe) > 0.01:
        if bal_owe > 0:
            lines.append(f"📤 *Остаток: мы должны {fmt_rub(bal_owe)}*")
        else:
            lines.append(f"⚠️ Переплата: мы вернули на {fmt_rub(-bal_owe)} больше")
    if abs(bal_owed) < 0.01 and abs(bal_owe) < 0.01:
        lines.append("✅ Расчёты закрыты.")

    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("client"))
async def on_client(message: Message):
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer(
            "Использование: `/client <имя | @username>`\n"
            "Например: `/client Стеф` или `/client @vasya_crypto`",
            parse_mode="Markdown")
        return
    name = parts[1].strip()
    c = get_client_stats(message.chat.id, name)
    if not c:
        await message.answer(f"Операций с «{name}» не нашёл.")
        return

    lines = [f"👤 *{c['display_name']}*  ({c['total_ops']} операций)\n"]

    # USDT-сделки
    if c["sell_count"] or c["buy_count"]:
        lines.append("📊 *USDT-сделки:*")
        if c["sell_count"]:
            lines.append(f"   📤 Продал ему: {c['sell_count']} раз, "
                         f"{fmt_usdt(c['sell_usdt'])} на {fmt_rub(c['sell_rub'])}")
            if c["avg_sell"]:
                lines.append(f"      ср. курс продажи {fmt_rate(c['avg_sell'])}")
        if c["buy_count"]:
            lines.append(f"   📥 Купил у него: {c['buy_count']} раз, "
                         f"{fmt_usdt(c['buy_usdt'])} на {fmt_rub(c['buy_rub'])}")
            if c["avg_buy"]:
                lines.append(f"      ср. курс закупки {fmt_rate(c['avg_buy'])}")
        if c["spread"] is not None:
            sign = "+" if c["spread"] >= 0 else ""
            emoji = "📈" if c["spread"] >= 0 else "📉"
            lines.append(f"   {emoji} Спред с ним: {sign}{c['spread']:.3f}₽/USDT")
        lines.append(f"   💰 Оборот: {fmt_rub(c['turnover_rub'])} / {fmt_usdt(c['turnover_usdt'])}")

    # Кэш
    if c["cash_in"] or c["cash_out"]:
        lines.append("\n💵 *Кэш:*")
        if c["cash_in"]:
            lines.append(f"   Принял от него: {fmt_rub(c['cash_in'])}")
        if c["cash_out"]:
            lines.append(f"   Выдал ему: {fmt_rub(c['cash_out'])}")

    # Долги
    if c["loan_out"] or c["loan_in"] or c["repay_in"] or c["repay_out"]:
        lines.append("\n💳 *Долги:*")
        if c["debt_balance"] > 0.01:
            lines.append(f"   📥 Должен нам сейчас: *{fmt_rub(c['debt_balance'])}*")
        elif c["debt_balance"] < -0.01:
            lines.append(f"   📤 Мы должны ему: *{fmt_rub(-c['debt_balance'])}*")
        else:
            lines.append("   ✅ По долгам в расчёте")

    if c["arb_count"]:
        lines.append(f"\n🔄 Участвовал в {c['arb_count']} арбитражных сделках")

    lines.append(f"\n_Операции {fmt_id(c['first_id'])}…{fmt_id(c['last_id'])}_")
    lines.append("_Вся история: /find " + c["display_name"] + "_")

    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("clients"))
async def on_clients(message: Message):
    parts = message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""

    clients = get_all_clients(message.chat.id)
    if not clients:
        await message.answer("USDT-сделок с клиентами ещё не было.")
        return

    # Сортировка
    sort_label = "по обороту"
    if arg in ("spread", "спред"):
        clients = [c for c in clients if c["spread"] is not None]
        clients.sort(key=lambda x: -(x["spread"] or 0))
        sort_label = "по спреду (выгодные сверху)"
    elif arg in ("ops", "частые", "count"):
        clients.sort(key=lambda x: -x["ops"])
        sort_label = "по числу сделок"
    # иначе уже отсортировано по обороту

    top = clients[:15]
    lines = [f"👥 *Клиенты ({len(clients)}) — {sort_label}:*\n"]
    for i, c in enumerate(top, 1):
        spread_str = ""
        if c["spread"] is not None:
            sign = "+" if c["spread"] >= 0 else ""
            spread_str = f", спред {sign}{c['spread']:.2f}"
        lines.append(
            f"{i}. *{c['display_name']}* — {fmt_rub(c['turnover_rub'])} "
            f"({c['ops']} сд.{spread_str})"
        )

    if len(clients) > 15:
        lines.append(f"\n_…и ещё {len(clients) - 15}_")

    lines.append("\n_Детали: `/client <имя>`_")
    lines.append("_Сортировка: `/clients spread` · `/clients частые`_")
    lines.append("_Таблицей: `/ct`_")

    await message.answer("\n".join(lines), parse_mode="Markdown")


def _fmt_money_short(n: float) -> str:
    """Сжатое представление: 1.5M, 729k, 950."""
    if n is None:
        return "—"
    n = float(n)
    if abs(n) >= 1_000_000:
        return f"{n/1_000_000:.1f}M"
    if abs(n) >= 1000:
        return f"{int(round(n/1000))}k"
    return str(int(round(n)))


def _fmt_spread_short(s) -> str:
    if s is None:
        return "  —"
    return f"+{s:.2f}" if s >= 0 else f"{s:.2f}"


@dp.message(Command("ct"))
@dp.message(Command("clients_table"))
async def on_clients_table(message: Message):
    """Таблица клиентов прямо в Telegram (моноширинная)."""
    parts = message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""

    clients = get_all_clients(message.chat.id)
    if not clients:
        await message.answer("USDT-сделок с клиентами ещё не было.")
        return

    sort_label = "по обороту"
    if arg in ("spread", "спред"):
        clients = [c for c in clients if c["spread"] is not None]
        clients.sort(key=lambda x: -(x["spread"] or 0))
        sort_label = "по спреду"
    elif arg in ("ops", "частые", "count"):
        clients.sort(key=lambda x: -x["ops"])
        sort_label = "по числу сделок"
    elif arg in ("usdt",):
        clients.sort(key=lambda x: -x["turnover_usdt"])
        sort_label = "по USDT-обороту"

    top = clients[:20]

    # Колонки: №  Клиент(15)  Сд(3)  Оборот(7)  Спред(5)
    header = f"{'№':<2}{'Клиент':<15} {'Сд':<3}{'Оборот':<7} {'Спред':<5}"
    sep = "─" * 36
    lines = [header, sep]
    for i, c in enumerate(top, 1):
        name = c["display_name"] or "—"
        if len(name) > 14:
            name = name[:13] + "…"
        idx = f"{i:<2}"
        lines.append(
            f"{idx}{name:<15} {c['ops']:<3}"
            f"{_fmt_money_short(c['turnover_rub']):<7} "
            f"{_fmt_spread_short(c['spread'])}"
        )

    table = "```\n" + "\n".join(lines) + "\n```"

    total_turnover = sum(c["turnover_rub"] for c in clients)
    total_usdt = sum(c["turnover_usdt"] for c in clients)

    footer = [
        f"\n📊 *Таблица клиентов* ({len(clients)}) — {sort_label}",
        table,
        f"💰 Общий оборот: *{fmt_rub(total_turnover)}* / *{fmt_usdt(total_usdt)}*",
    ]
    if len(clients) > 20:
        footer.append(f"_Показаны топ-20 из {len(clients)}_")
    footer.append(
        "\n_Сортировка:_ `/ct spread` · `/ct частые` · `/ct usdt`\n"
        "_Полная Excel-таблица:_ `/export`"
    )

    await message.answer("\n".join(footer), parse_mode="Markdown")


@dp.message(Command("day"))
async def on_day(message: Message):
    agg = get_period_report(message.chat.id, "day")
    await message.answer(format_period_report(agg), parse_mode="Markdown")


@dp.message(Command("week"))
async def on_week(message: Message):
    agg = get_period_report(message.chat.id, "week")
    await message.answer(format_period_report(agg), parse_mode="Markdown")


@dp.message(Command("month"))
async def on_month(message: Message):
    agg = get_period_report(message.chat.id, "month")
    await message.answer(format_period_report(agg), parse_mode="Markdown")


@dp.message(Command("export"))
async def on_export(message: Message):
    """Выгрузка операций в Excel. /export | /export day|week|month"""
    parts = message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""
    period = None
    label = "все операции"
    if arg in ("day", "день", "сегодня"):
        period = "day"; label = "за сегодня"
    elif arg in ("week", "неделя"):
        period = "week"; label = "за неделю"
    elif arg in ("month", "месяц"):
        period = "month"; label = "за месяц"
    elif arg:
        await message.answer(
            "Использование:\n"
            "`/export` — все операции\n"
            "`/export day` — за сегодня\n"
            "`/export week` — за неделю\n"
            "`/export month` — за месяц",
            parse_mode="Markdown")
        return

    await message.answer("📊 Готовлю Excel-файл…")
    try:
        path = build_excel_export(message.chat.id, period=period)
    except Exception as e:
        log.exception("Ошибка экспорта")
        await message.answer(f"Не удалось собрать файл: {e}")
        return

    try:
        doc = FSInputFile(path, filename=Path(path).name)
        await message.answer_document(
            doc, caption=f"📊 Выгрузка ({label}). Открывается в Excel / Google Таблицах.")
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


@dp.message(Command("backup"))
async def on_backup(message: Message):
    """Шлёт файл базы данных — резервная копия. Храни в надёжном месте."""
    if not DB_PATH.exists():
        await message.answer("База ещё пуста — нечего сохранять.")
        return
    await message.answer("💾 Готовлю резервную копию базы…")
    try:
        stamp = datetime.now(LOCAL_TZ).strftime("%Y%m%d_%H%M")
        doc = FSInputFile(str(DB_PATH), filename=f"trades_backup_{stamp}.db")
        await message.answer_document(
            doc,
            caption=(
                "💾 *Резервная копия базы*\n\n"
                "Сохрани этот файл. Если что-то случится с ботом или сервером — "
                "пришли мне его обратно, и я подскажу как восстановить.\n\n"
                "_Делай бэкап раз в несколько дней или после большого объёма операций._"
            ),
            parse_mode="Markdown")
        # Запоминаем момент бэкапа — чтобы напоминалка не дёргала зря
        set_chat_setting(message.chat.id, last_backup_ts=datetime.now(LOCAL_TZ).isoformat())
    except Exception as e:
        log.exception("Ошибка бэкапа")
        await message.answer(f"Не удалось отправить базу: {e}")


@dp.message(Command("reminders"))
async def on_reminders(message: Message):
    """Включить/выключить автонапоминания (бэкап + оплата Railway)."""
    parts = message.text.split(maxsplit=1)
    s = get_chat_settings(message.chat.id)
    if len(parts) < 2:
        status = "включены ✅" if s.get("reminders_on") else "выключены ❌"
        pay = s.get("pay_date")
        pay_line = f"\n💳 Дата оплаты Railway: *{pay}*" if pay else "\n💳 Дата оплаты Railway не задана (см. /setpaydate)"
        last_bk = s.get("last_backup_ts")
        bk_line = ""
        if last_bk:
            try:
                d = datetime.fromisoformat(last_bk)
                bk_line = f"\n💾 Последний бэкап: {d.strftime('%d.%m.%Y %H:%M')}"
            except (ValueError, TypeError):
                pass
        await message.answer(
            f"🔔 *Напоминания {status}*\n"
            f"{pay_line}{bk_line}\n\n"
            "Бот сам напомнит:\n"
            "• 💾 сделать бэкап, если давно не делал (раз в ~4 дня)\n"
            "• 💳 про оплату Railway за несколько дней до даты\n\n"
            "Управление:\n"
            "`/reminders on` — включить\n"
            "`/reminders off` — выключить\n"
            "`/setpaydate 21.06.2026` — задать дату оплаты Railway",
            parse_mode="Markdown")
        return
    arg = parts[1].strip().lower()
    if arg in ("on", "вкл", "включить"):
        set_chat_setting(message.chat.id, reminders_on=1)
        await message.answer(
            "🔔 Напоминания включены.\n\n"
            "Буду писать сюда про бэкапы и оплату Railway.\n"
            "Чтобы я напоминал и про оплату — задай дату: `/setpaydate 21.06.2026`",
            parse_mode="Markdown")
    elif arg in ("off", "выкл", "выключить"):
        set_chat_setting(message.chat.id, reminders_on=0)
        await message.answer("🔕 Напоминания выключены.")
    else:
        await message.answer("Используй `/reminders on` или `/reminders off`.", parse_mode="Markdown")


@dp.message(Command("setpaydate"))
async def on_setpaydate(message: Message):
    """Задать дату оплаты Railway: /setpaydate 21.06.2026"""
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer(
            "Использование: `/setpaydate ДД.ММ.ГГГГ`\n"
            "Например: `/setpaydate 21.06.2026`\n\n"
            "Посмотреть дату можно в Railway вверху справа "
            "(«24 days left» и т.п.). Прибавь это число дней к сегодня.",
            parse_mode="Markdown")
        return
    raw = parts[1].strip()
    # Принимаем ДД.ММ.ГГГГ или ГГГГ-ММ-ДД
    parsed = None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(raw, fmt)
            break
        except ValueError:
            continue
    if not parsed:
        await message.answer("Не понял дату. Пример: `/setpaydate 21.06.2026`", parse_mode="Markdown")
        return
    iso = parsed.strftime("%Y-%m-%d")
    set_chat_setting(message.chat.id, pay_date=iso, reminders_on=1)
    await message.answer(
        f"💳 Дата оплаты Railway: *{parsed.strftime('%d.%m.%Y')}*\n"
        "Напомню за 7, 3 и 1 день до неё (и в сам день).\n"
        "_Напоминания заодно включил._",
        parse_mode="Markdown")


@dp.message(Command("pending"))
async def on_pending(message: Message):
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'pending' "
            "ORDER BY id",
            (message.chat.id,)).fetchall()
    if not rows:
        await message.answer("Нет операций, ждущих подтверждения.")
        return
    lines = [f"⏳ *Ждут подтверждения ({len(rows)}):*\n"]
    for r in rows:
        lines.append(fmt_tx_short(r))
    lines.append("\nИспользуй кнопки под нужным сообщением, или /pending_cancel чтобы удалить все.")
    await message.answer("\n".join(lines), parse_mode="Markdown")


@dp.message(Command("pending_cancel"))
async def on_pending_cancel(message: Message):
    with db() as conn:
        n = conn.execute(
            "SELECT COUNT(*) AS c FROM transactions WHERE chat_id = ? AND status = 'pending'",
            (message.chat.id,)).fetchone()["c"]
        conn.execute(
            "DELETE FROM transactions WHERE chat_id = ? AND status = 'pending'",
            (message.chat.id,))
    await message.answer(f"🗑 Удалил {n} неподтверждённых операций.")


@dp.message(Command("find"))
async def on_find(message: Message):
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer(
            "Использование: `/find <ID | имя | сумма>`\n\n"
            "`/find #142` — найти операцию по ID\n"
            "`/find Адахан` — все операции с этим контрагентом\n"
            "`/find 96000` — все операции на эту сумму",
            parse_mode="Markdown")
        return
    query = parts[1].strip()

    id_match = re.match(r"^#?(\d+)$", query.replace(" ", ""))
    if id_match:
        tx_id = int(id_match.group(1))
        with db() as conn:
            row = conn.execute("SELECT * FROM transactions WHERE id = ?", (tx_id,)).fetchone()
        if not row:
            await message.answer(f"Операция #{tx_id:03d} не найдена.")
            return
        same_chat = row["chat_id"] == message.chat.id
        note = "" if same_chat else f"\n_(из другого чата, ID={row['chat_id']})_"
        await message.answer(f"🔍 *Найдено:*\n\n{fmt_tx_short(row)}{note}", parse_mode="Markdown")
        return

    try:
        amount = parse_number(query)
        if amount >= 100:
            with db() as conn:
                rows = conn.execute(
                    "SELECT * FROM transactions WHERE chat_id = ? "
                    "AND ABS(rubles - ?) < 0.5 AND status = 'confirmed' ORDER BY id DESC LIMIT 15",
                    (message.chat.id, amount)).fetchall()
            if rows:
                out = [f"🔍 *Найдено {len(rows)} операций на {fmt_rub(amount)}:*\n"]
                out.extend(fmt_tx_short(r) for r in rows)
                await message.answer("\n".join(out), parse_mode="Markdown")
                return
    except (ValueError, TypeError):
        pass

    pattern = f"%{query.lower()}%"
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
            "AND (pylower(counterparty) LIKE ? OR pylower(partner_in) LIKE ? OR pylower(partner_out) LIKE ?) "
            "ORDER BY id DESC LIMIT 15",
            (message.chat.id, pattern, pattern, pattern)).fetchall()
    if not rows:
        await message.answer(f"По запросу «{query}» ничего не нашёл.")
        return
    out = [f"🔍 *Найдено {len(rows)} операций по «{query}»:*\n"]
    out.extend(fmt_tx_short(r) for r in rows)
    await message.answer("\n".join(out), parse_mode="Markdown")


@dp.message(Command("undo"))
async def on_undo(message: Message):
    chat_id = message.chat.id
    parts = message.text.split(maxsplit=1)

    # /undo #142 — отменить конкретную операцию
    target_id = None
    if len(parts) > 1:
        m = re.match(r"^#?(\d+)$", parts[1].strip().replace(" ", ""))
        if m:
            target_id = int(m.group(1))
        else:
            await message.answer(
                "Использование:\n"
                "`/undo` — отменить последнюю операцию\n"
                "`/undo #142` — отменить операцию по номеру",
                parse_mode="Markdown")
            return

    with db() as conn:
        if target_id is not None:
            row = conn.execute(
                "SELECT * FROM transactions WHERE chat_id = ? AND id = ? AND status = 'confirmed'",
                (chat_id, target_id)).fetchone()
            if not row:
                await message.answer(
                    f"Операцию {fmt_id(target_id)} не нашёл "
                    "(возможно, она в другом чате или уже отменена).")
                return
        else:
            row = conn.execute(
                "SELECT * FROM transactions WHERE chat_id = ? AND status = 'confirmed' "
                "ORDER BY id DESC LIMIT 1",
                (chat_id,)).fetchone()
            if not row:
                await message.answer("Отменять нечего.")
                return
        conn.execute("DELETE FROM transactions WHERE id = ?", (row["id"],))

    reverse_balance(chat_id, row)
    s = get_state(chat_id)
    await message.answer(
        f"↩️ Отменил {fmt_id(row['id'])}: `{row['raw_text']}`\n"
        f"💰 Касса: *{fmt_rub(s['ruble_balance'])}*  💵 *{fmt_usdt(s['usdt_balance'])}*",
        parse_mode="Markdown")


@dp.message(Command("reset"))
async def on_reset(message: Message):
    with db() as conn:
        conn.execute("DELETE FROM transactions WHERE chat_id = ?", (message.chat.id,))
        conn.execute(
            "UPDATE state SET ruble_balance=0, usdt_balance=0, extra_rubles=0 WHERE chat_id = ?",
            (message.chat.id,))
    await message.answer("🗑 Всё обнулил.")


@dp.message(Command("setcash"))
async def on_setcash(message: Message):
    parts = message.text.split()
    if len(parts) < 2:
        await message.answer(
            "Использование: `/setcash <рубли> [USDT]`\nНапример: `/setcash 4348000 1500`",
            parse_mode="Markdown")
        return
    try:
        rub = parse_number(parts[1])
        usdt = parse_number(parts[2]) if len(parts) > 2 else 0
    except ValueError:
        await message.answer("Не понял числа. Пример: `/setcash 4348000 1500`", parse_mode="Markdown")
        return
    with db() as conn:
        conn.execute(
            "INSERT INTO state (chat_id, ruble_balance, usdt_balance) VALUES (?, ?, ?) "
            "ON CONFLICT(chat_id) DO UPDATE SET ruble_balance = excluded.ruble_balance, "
            "usdt_balance = excluded.usdt_balance",
            (message.chat.id, rub, usdt))
    await message.answer(f"✅ Касса: {fmt_rub(rub)}\n💵 Кошелёк: {fmt_usdt(usdt)}")


@dp.message(Command("extra"))
async def on_extra(message: Message):
    parts = message.text.split()
    if len(parts) < 2:
        await message.answer("Использование: `/extra <сумма>`", parse_mode="Markdown")
        return
    try:
        val = parse_number(parts[1])
    except ValueError:
        await message.answer("Не понял число.")
        return
    with db() as conn:
        conn.execute("INSERT OR IGNORE INTO state (chat_id) VALUES (?)", (message.chat.id,))
        conn.execute("UPDATE state SET extra_rubles = ? WHERE chat_id = ?", (val, message.chat.id))
    await message.answer(f"♻️ Лишние: {fmt_rub(val)}")


# ────────────────── ОСНОВНОЙ ОБРАБОТЧИК СООБЩЕНИЙ ──────────────────
@dp.message(F.text)
async def on_text(message: Message):
    chat_id = message.chat.id
    settings = get_chat_settings(chat_id)

    # 1. Сначала пробуем распознать арбитраж
    arb = parse_arbitrage_message(message.text)
    if arb:
        await handle_arbitrage(message, arb, settings)
        return

    # 2. Иначе — строка за строкой
    parsed_items = []
    for line in message.text.split("\n"):
        p = parse_message_line(line)
        if p:
            parsed_items.append((p, line.strip()))

    if not parsed_items:
        low = message.text.lower().strip()
        trade_triggers = ("продал", "продала", "продали", "купил", "купила",
                          "купили", "откупил", "откупили")
        cash_loan_triggers = ("выдал", "выдала", "выдали", "отдал", "отдали",
                    "принял", "приняла", "приняли", "забрал", "забрали",
                    "занял", "заняла", "заняли", "долг", "вернул",
                    "вернула", "вернули", "погасил", "погасили", "дал")
        has_number = bool(re.search(r"\d", low))
        has_op = bool(re.search(r"[*x×/÷:]", low))

        if any(t in low for t in trade_triggers):
            # Сделка, но не распозналась
            hint = "Не понял сделку. "
            if not has_number:
                hint += "Не вижу чисел.\n"
            elif not has_op:
                hint += "Похоже нет знака `*` или `/`.\n_`*` — слева USDT, `/` — слева рубли._\n"
            else:
                hint += "Проверь формат:\n"
            await message.reply(
                hint +
                "`Продал Гиге 100*76`  (100 USDT по 76)\n"
                "`Купил у Васи 75000/74`  (на 75000₽ по 74)",
                parse_mode="Markdown")
        elif any(t in low for t in cash_loan_triggers):
            await message.reply(
                "Не понял операцию. Примеры:\n"
                "`Выдал Владу 72600`\n"
                "`Принял от Адахана 96000`\n"
                "`Дал в долг Михаилу 90000`\n"
                "`Михаил вернул 50000`",
                parse_mode="Markdown")
        return

    await handle_regular_ops(message, parsed_items, settings)


async def handle_arbitrage(message: Message, arb: dict, settings: dict):
    chat_id = message.chat.id
    needs_confirm = should_confirm(arb, settings)
    status = "pending" if needs_confirm else "confirmed"

    # raw_text — оригинал сообщения для трассировки
    raw = message.text.strip()
    counterparty = f"{arb['partner_in']}→{arb['partner_out']}"

    tid = save_tx(
        chat_id, "arb", counterparty, arb["rub_amount"],
        usdt=arb["profit"],
        rate=0,
        raw=raw,
        status=status,
        usdt_in=arb["usdt_in"],
        usdt_out=arb["usdt_out"],
        partner_in=arb["partner_in"],
        partner_out=arb["partner_out"],
    )

    initial = get_state(chat_id)

    lines = [
        f"🔄 *Арбитраж* {fmt_id(tid)}{' ⏳' if needs_confirm else ''}",
        "",
        f"📥 От *{arb['partner_in']}*: +{fmt_usdt(arb['usdt_in'])} (по курсу {fmt_rate(arb['rate_in'])})",
        f"📤 К *{arb['partner_out']}*: −{fmt_usdt(arb['usdt_out'])} (по курсу {fmt_rate(arb['rate_out'])})",
        f"💰 Через кассу прошло: {fmt_rub(arb['rub_amount'])} _(но касса не меняется)_",
        "",
        f"💵 *Профит: +{fmt_usdt(arb['profit'])}*",
    ]

    if needs_confirm:
        lines.append("\n⚠️ Проверь и подтверди:")
        await message.reply("\n".join(lines), parse_mode="Markdown", reply_markup=confirm_kb(tid))
    else:
        apply_balance(chat_id, {
            "type": "arb",
            "usdt": arb["profit"],
            "rubles": 0,
        })
        final = get_state(chat_id)
        old_arb = initial.get("arb_profit_usdt", 0) or 0
        new_arb = final.get("arb_profit_usdt", 0) or 0
        lines.append(
            f"\n📊 *Касса и кошелёк USDT не изменились.*\n"
            f"💼 Доход с арбитража: {fmt_usdt(old_arb)} +{fmt_usdt(arb['profit'])} "
            f"= *{fmt_usdt(new_arb)}*"
        )
        debt_status = format_debt_status(get_debts(chat_id))
        if debt_status:
            lines.append(debt_status)
        await message.reply("\n".join(lines), parse_mode="Markdown")


async def handle_regular_ops(message: Message, parsed_items: list, settings: dict):
    chat_id = message.chat.id

    # Решаем, есть ли среди операций такие, что требуют подтверждения
    needs_confirm_items = [p for p, _ in parsed_items if should_confirm(p, settings)]
    immediate_items = [(p, raw) for p, raw in parsed_items if not should_confirm(p, settings)]

    # ВАРИАНТ 1: смешанный режим (часть нужна подтверждения, часть — нет)
    # Чтобы не запутать пользователя — всё подтверждать вместе, если хоть одна нужна.
    if needs_confirm_items:
        # Сохраняем ВСЕ как pending, общий batch_id = id первой
        saved = []
        first_id = None
        for p, raw in parsed_items:
            t = p["type"]
            cp = p["counterparty"]
            if t in ("sell", "buy"):
                tid = save_tx(chat_id, t, cp, p["rubles"], usdt=p["usdt"], rate=p["rate"],
                              raw=raw, doer=p.get("doer"), status="pending")
            else:
                tid = save_tx(chat_id, t, cp, p["amount"],
                              raw=raw, doer=p.get("doer"), status="pending")
            if first_id is None:
                first_id = tid
            saved.append((tid, p))

        # Привязываем все к batch_id первой операции
        with db() as conn:
            ids = [s[0] for s in saved]
            placeholders = ",".join("?" * len(ids))
            conn.execute(
                f"UPDATE transactions SET batch_id = ? WHERE id IN ({placeholders})",
                (first_id, *ids))

        # Сборка превью
        if len(saved) == 1:
            tid, p = saved[0]
            t = p["type"]
            if t in ("sell", "buy"):
                verb = "Продал" if t == "sell" else "Купил у"
                doer_str = f" ({p['doer']})" if p.get("doer") else ""
                reply = [
                    f"{fmt_id(tid)} ⏳ {verb} *{p['counterparty']}*{doer_str}",
                    f"   {fmt_usdt(p['usdt'])} × {fmt_rate(p['rate'])} = {fmt_rub(p['rubles'])}",
                    "",
                    "⚠️ *Проверь и подтверди:*",
                ]
                if t == "sell":
                    reply.append(f"📥 Ты получил {fmt_rub(p['rubles'])}?")
                    reply.append(f"📤 Ты отправил {fmt_usdt(p['usdt'])}?")
                else:
                    reply.append(f"📥 Ты получил {fmt_usdt(p['usdt'])}?")
                    reply.append(f"📤 Ты отправил {fmt_rub(p['rubles'])}?")
                await message.reply("\n".join(reply), parse_mode="Markdown",
                                    reply_markup=confirm_kb(first_id))
            else:
                sign = "+" if t == "cash_in" else "−"
                verb = "Принял" if t == "cash_in" else "Выдал"
                doer_str = f" ({p['doer']})" if p.get("doer") else ""
                reply = [
                    f"{fmt_id(tid)} ⏳ {verb} *{p['counterparty']}*{doer_str}",
                    f"   {sign}{fmt_rub(p['amount'])}",
                    "",
                    "⚠️ *Подтверди:*",
                ]
                await message.reply("\n".join(reply), parse_mode="Markdown",
                                    reply_markup=confirm_kb(first_id))
        else:
            lines = [f"⏳ *Распарсил {len(saved)} операций — подтверди все:*\n"]
            for tid, p in saved:
                t = p["type"]
                if t == "sell":
                    lines.append(f"{fmt_id(tid)} 📤 Продал {p['counterparty']}: +{fmt_rub(p['rubles'])} / −{fmt_usdt(p['usdt'])}")
                elif t == "buy":
                    lines.append(f"{fmt_id(tid)} 📥 Купил у {p['counterparty']}: −{fmt_rub(p['rubles'])} / +{fmt_usdt(p['usdt'])}")
                elif t == "cash_in":
                    lines.append(f"{fmt_id(tid)} 💰 Принял {p['counterparty']}: +{fmt_rub(p['amount'])}")
                elif t == "cash_out":
                    lines.append(f"{fmt_id(tid)} 💸 Выдал {p['counterparty']}: −{fmt_rub(p['amount'])}")
            await message.reply("\n".join(lines), parse_mode="Markdown",
                                reply_markup=confirm_kb(first_id))
        return

    # ВАРИАНТ 2: ничего не требует подтверждения — сразу записываем
    initial = get_state(chat_id)
    saved = []
    for p, raw in parsed_items:
        t = p["type"]
        cp = p["counterparty"]
        if t in ("sell", "buy"):
            tid = save_tx(chat_id, t, cp, p["rubles"], usdt=p["usdt"], rate=p["rate"],
                          raw=raw, doer=p.get("doer"), status="confirmed")
            apply_balance(chat_id, {"type": t, "usdt": p["usdt"], "rubles": p["rubles"]})
        else:
            tid = save_tx(chat_id, t, cp, p["amount"],
                          raw=raw, doer=p.get("doer"), status="confirmed")
            apply_balance(chat_id, {"type": t, "usdt": 0, "rubles": p["amount"]})
        saved.append((tid, p))

    final = get_state(chat_id)

    if len(saved) == 1:
        tid, p = saved[0]
        t = p["type"]
        if t in ("sell", "buy"):
            verb = "Продал" if t == "sell" else "Купил у"
            op_sign = "+" if t == "sell" else "−"
            doer_str = f" ({p['doer']})" if p.get("doer") else ""
            d = get_debts(chat_id)
            debt_reminder = format_debt_status(d)

            # Подсветка плохой сделки: сравниваем курс этой сделки со средними
            # ДО неё (по текущему циклу). Если sell ниже средней закупки —
            # продаём в убыток. Если buy выше средней продажи — закупаемся дороже,
            # чем планируем продавать.
            trade_warning = ""
            prev_rates = get_avg_rates_excluding(chat_id, exclude_id=tid, period="cycle")
            if t == "sell" and prev_rates["avg_buy"] is not None:
                if p["rate"] < prev_rates["avg_buy"]:
                    diff = prev_rates["avg_buy"] - p["rate"]
                    loss = diff * p["usdt"]
                    trade_warning = (
                        f"\n\n⚠️ *Курс {fmt_rate(p['rate'])} ниже средней закупки "
                        f"{fmt_rate(prev_rates['avg_buy'])} — продал в убыток*\n"
                        f"   −{fmt_rate(diff)}₽/USDT  (≈ −{fmt_rub(loss)} на этой сделке)"
                    )
            elif t == "buy" and prev_rates["avg_sell"] is not None:
                if p["rate"] > prev_rates["avg_sell"]:
                    diff = p["rate"] - prev_rates["avg_sell"]
                    loss = diff * p["usdt"]
                    trade_warning = (
                        f"\n\n⚠️ *Курс {fmt_rate(p['rate'])} выше средней продажи "
                        f"{fmt_rate(prev_rates['avg_sell'])} — закупил дороже планируемой продажи*\n"
                        f"   +{fmt_rate(diff)}₽/USDT  (≈ −{fmt_rub(loss)} закладываешь в убыток)"
                    )

            await message.reply(
                f"{fmt_id(tid)} ✅ {verb} *{p['counterparty']}*{doer_str}\n"
                f"   {fmt_usdt(p['usdt'])} × {fmt_rate(p['rate'])} = {fmt_rub(p['rubles'])}\n\n"
                f"💰 Касса: {fmt_rub(initial['ruble_balance'])} {op_sign} {fmt_rub(p['rubles'])} = *{fmt_rub(final['ruble_balance'])}*\n"
                f"💵 Кошелёк: *{fmt_usdt(final['usdt_balance'])}*"
                f"{trade_warning}"
                f"{debt_reminder}",
                parse_mode="Markdown")
        else:
            doer_str = f" ({p['doer']})" if p.get("doer") else ""
            if t == "loan_out":
                verb_full = f"Дал в долг *{p['counterparty']}*{doer_str}"
                op_sign = "−"
                sign = "−"
            elif t == "loan_in":
                verb_full = f"Занял у *{p['counterparty']}*{doer_str}"
                op_sign = "+"
                sign = "+"
            elif t == "debt_repay_in":
                verb_full = f"*{p['counterparty']}*{doer_str} вернул долг"
                op_sign = "+"
                sign = "+"
            elif t == "debt_repay_out":
                verb_full = f"Вернул долг *{p['counterparty']}*{doer_str}"
                op_sign = "−"
                sign = "−"
            else:
                verb = "Принял" if t == "cash_in" else "Выдал"
                verb_full = f"{verb} *{p['counterparty']}*{doer_str}"
                op_sign = "+" if t == "cash_in" else "−"
                sign = op_sign

            reply_lines = [
                f"{fmt_id(tid)} ✅ {verb_full}",
                f"   {sign}{fmt_rub(p['amount'])}",
                "",
                f"💰 Касса: {fmt_rub(initial['ruble_balance'])} {op_sign} {fmt_rub(p['amount'])} = *{fmt_rub(final['ruble_balance'])}*",
            ]
            # Для долговых операций — показываем актуальное состояние долгов
            if t in ("loan_out", "loan_in", "debt_repay_in", "debt_repay_out"):
                d = get_debts(chat_id)
                debt_line = format_debt_status(d, focus_name=p["counterparty"])
                if debt_line:
                    reply_lines.append(debt_line)
            await message.reply("\n".join(reply_lines), parse_mode="Markdown")
    else:
        lines = [f"✅ *Записано {len(saved)} операций:*\n"]
        d_rub = 0.0
        d_usdt = 0.0
        has_debt_ops = False
        for tid, p in saved:
            t = p["type"]
            if t == "sell":
                lines.append(f"{fmt_id(tid)} 📤 Продал {p['counterparty']}: +{fmt_rub(p['rubles'])} / −{fmt_usdt(p['usdt'])}")
                d_rub += p["rubles"]; d_usdt -= p["usdt"]
            elif t == "buy":
                lines.append(f"{fmt_id(tid)} 📥 Купил у {p['counterparty']}: −{fmt_rub(p['rubles'])} / +{fmt_usdt(p['usdt'])}")
                d_rub -= p["rubles"]; d_usdt += p["usdt"]
            elif t == "cash_in":
                lines.append(f"{fmt_id(tid)} 💰 Принял {p['counterparty']}: +{fmt_rub(p['amount'])}")
                d_rub += p["amount"]
            elif t == "cash_out":
                lines.append(f"{fmt_id(tid)} 💸 Выдал {p['counterparty']}: −{fmt_rub(p['amount'])}")
                d_rub -= p["amount"]
            elif t == "loan_out":
                lines.append(f"{fmt_id(tid)} 📤💳 Дал в долг {p['counterparty']}: −{fmt_rub(p['amount'])}")
                d_rub -= p["amount"]; has_debt_ops = True
            elif t == "loan_in":
                lines.append(f"{fmt_id(tid)} 📥💳 Занял у {p['counterparty']}: +{fmt_rub(p['amount'])}")
                d_rub += p["amount"]; has_debt_ops = True
            elif t == "debt_repay_in":
                lines.append(f"{fmt_id(tid)} ↩️💰 {p['counterparty']} вернул: +{fmt_rub(p['amount'])}")
                d_rub += p["amount"]; has_debt_ops = True
            elif t == "debt_repay_out":
                lines.append(f"{fmt_id(tid)} ↩️💸 Вернул {p['counterparty']}: −{fmt_rub(p['amount'])}")
                d_rub -= p["amount"]; has_debt_ops = True
        lines.append(f"\n📊 Итого по ₽: {('+' if d_rub >= 0 else '−')}{fmt_rub(abs(d_rub))}")
        if abs(d_usdt) > 1e-6:
            lines.append(f"📊 Итого по USDT: {('+' if d_usdt >= 0 else '−')}{fmt_usdt(abs(d_usdt))}")
        lines.append(f"\n💰 Касса: {fmt_rub(initial['ruble_balance'])} → *{fmt_rub(final['ruble_balance'])}*")
        if abs(d_usdt) > 1e-6:
            lines.append(f"💵 Кошелёк: *{fmt_usdt(final['usdt_balance'])}*")
        if has_debt_ops:
            debt_status = format_debt_status(get_debts(chat_id))
            if debt_status:
                lines.append(debt_status)
        await message.reply("\n".join(lines), parse_mode="Markdown")


# ────────────────── ОБРАБОТЧИКИ КНОПОК ──────────────────
@dp.callback_query(F.data.startswith("c:"))
async def on_confirm_button(cq: CallbackQuery):
    try:
        batch_id = int(cq.data[2:])
    except ValueError:
        await cq.answer("Ошибка кнопки.", show_alert=True)
        return

    chat_id = cq.message.chat.id
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND batch_id = ? AND status = 'pending'",
            (chat_id, batch_id)).fetchall()
        # Если batch_id не выставлен (одиночная операция), пробуем по id
        if not rows:
            rows = conn.execute(
                "SELECT * FROM transactions WHERE chat_id = ? AND id = ? AND status = 'pending'",
                (chat_id, batch_id)).fetchall()
        if not rows:
            await cq.answer("Уже подтверждено или отменено.", show_alert=True)
            try:
                await cq.message.edit_reply_markup(reply_markup=None)
            except Exception:
                pass
            return

        # Подтверждаем все
        ids = [r["id"] for r in rows]
        placeholders = ",".join("?" * len(ids))
        conn.execute(
            f"UPDATE transactions SET status = 'confirmed' WHERE id IN ({placeholders})",
            ids)

    # Применяем балансы
    for r in rows:
        apply_balance(chat_id, r)

    final = get_state(chat_id)
    n = len(rows)
    if n == 1:
        r = rows[0]
        confirm_line = f"\n\n✅ *Подтверждено* в {datetime.utcnow().strftime('%H:%M UTC')}"
    else:
        confirm_line = f"\n\n✅ *Подтверждено {n} операций* в {datetime.utcnow().strftime('%H:%M UTC')}"
    confirm_line += f"\n💰 Касса: *{fmt_rub(final['ruble_balance'])}*"
    confirm_line += f"\n💵 Кошелёк: *{fmt_usdt(final['usdt_balance'])}*"
    arb_total = final.get("arb_profit_usdt", 0) or 0
    if arb_total:
        confirm_line += f"\n💼 Доход с арбитража: *{fmt_usdt(arb_total)}*"

    # Подсветка плохих сделок среди подтверждённых
    bad_trades = []
    for r in rows:
        t = r["type"]
        if t not in ("sell", "buy"):
            continue
        prev = get_avg_rates_excluding(chat_id, exclude_id=r["id"], period="cycle")
        rate = r["rate"] or 0
        usdt_amt = r["usdt"] or 0
        if t == "sell" and prev["avg_buy"] is not None and rate < prev["avg_buy"]:
            diff = prev["avg_buy"] - rate
            loss = diff * usdt_amt
            bad_trades.append(
                f"⚠️ {fmt_id(r['id'])} продал по {fmt_rate(rate)} < ср. закупки "
                f"{fmt_rate(prev['avg_buy'])} (≈ −{fmt_rub(loss)})"
            )
        elif t == "buy" and prev["avg_sell"] is not None and rate > prev["avg_sell"]:
            diff = rate - prev["avg_sell"]
            loss = diff * usdt_amt
            bad_trades.append(
                f"⚠️ {fmt_id(r['id'])} купил по {fmt_rate(rate)} > ср. продажи "
                f"{fmt_rate(prev['avg_sell'])} (≈ −{fmt_rub(loss)})"
            )
    if bad_trades:
        confirm_line += "\n\n" + "\n".join(bad_trades)

    debt_status = format_debt_status(get_debts(chat_id))
    if debt_status:
        confirm_line += debt_status

    try:
        new_text = (cq.message.text or "") + confirm_line
        await cq.message.edit_text(new_text, parse_mode="Markdown", reply_markup=None)
    except Exception as e:
        log.warning(f"edit_text failed: {e}")
        await cq.message.answer(f"✅ Подтверждено {n} операций.{confirm_line}", parse_mode="Markdown")
    await cq.answer("Записано!")


@dp.callback_query(F.data.startswith("x:"))
async def on_cancel_button(cq: CallbackQuery):
    try:
        batch_id = int(cq.data[2:])
    except ValueError:
        await cq.answer("Ошибка кнопки.", show_alert=True)
        return

    chat_id = cq.message.chat.id
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE chat_id = ? AND batch_id = ? AND status = 'pending'",
            (chat_id, batch_id)).fetchall()
        if not rows:
            rows = conn.execute(
                "SELECT * FROM transactions WHERE chat_id = ? AND id = ? AND status = 'pending'",
                (chat_id, batch_id)).fetchall()
        if not rows:
            await cq.answer("Уже подтверждено или отменено.", show_alert=True)
            try:
                await cq.message.edit_reply_markup(reply_markup=None)
            except Exception:
                pass
            return

        ids = [r["id"] for r in rows]
        placeholders = ",".join("?" * len(ids))
        conn.execute(f"DELETE FROM transactions WHERE id IN ({placeholders})", ids)

    n = len(rows)
    try:
        new_text = (cq.message.text or "") + f"\n\n❌ *Отменено* ({n} оп.)"
        await cq.message.edit_text(new_text, parse_mode="Markdown", reply_markup=None)
    except Exception:
        await cq.message.answer(f"❌ Отменено {n} операций.", parse_mode="Markdown")
    await cq.answer("Отменено")


# ────────────────── ЗАПУСК ──────────────────
async def reminders_loop():
    """
    Фоновая задача: раз в час проверяет все чаты с включёнными напоминаниями
    и шлёт сообщения о бэкапе и оплате Railway, если пришло время.

    Тихие часы: не шлю с 22:00 до 09:00 по локальному времени.
    Анти-спам: один тип напоминания — не чаще раза в сутки.
    """
    # Небольшая задержка чтобы бот успел стартануть
    await asyncio.sleep(30)
    log.info("Reminders loop started.")
    while True:
        try:
            await _check_reminders_once()
        except Exception:
            log.exception("Reminders loop error")
        # Раз в час
        await asyncio.sleep(3600)


async def _check_reminders_once():
    now_local = datetime.now(LOCAL_TZ)
    if now_local.hour < 9 or now_local.hour >= 22:
        return  # тихие часы
    today_str = now_local.strftime("%Y-%m-%d")

    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM chat_settings WHERE reminders_on = 1").fetchall()

    for r in rows:
        chat_id = r["chat_id"]
        try:
            await _maybe_remind_backup(chat_id, r, now_local, today_str)
            await _maybe_remind_pay(chat_id, r, now_local, today_str)
        except Exception:
            log.exception(f"Reminder send error for chat {chat_id}")


async def _maybe_remind_backup(chat_id: int, settings, now_local, today_str: str):
    """Напоминание о бэкапе. Раз в 4 дня (по умолчанию), не чаще 1 в день."""
    interval_days = 4
    last_remind = settings["last_backup_remind"] if "last_backup_remind" in settings.keys() else None
    if last_remind == today_str:
        return  # уже напоминал сегодня

    last_backup = settings["last_backup_ts"] if "last_backup_ts" in settings.keys() else None
    days_since_backup = None
    if last_backup:
        try:
            d = datetime.fromisoformat(last_backup)
            if d.tzinfo is None:
                d = d.replace(tzinfo=LOCAL_TZ)
            days_since_backup = (now_local - d).days
        except (ValueError, TypeError):
            pass

    # Если бэкапа не было совсем — напомним через 2 дня после включения reminders
    # Если был — через interval_days
    should_remind = False
    if last_backup is None:
        # Бэкап ещё ни разу не делали — напомним один раз
        should_remind = True
        text = (
            "💾 *Напоминание о резервной копии*\n\n"
            "Ты ещё ни разу не делал бэкап базы. Сейчас самое время:\n\n"
            "Напиши `/backup` — я пришлю файл базы в чат. Просто сохрани его.\n\n"
            "_Делай так раз в несколько дней — это защита от любых сюрпризов._"
        )
    elif days_since_backup is not None and days_since_backup >= interval_days:
        should_remind = True
        text = (
            f"💾 *Напоминание о резервной копии*\n\n"
            f"Последний бэкап был {days_since_backup} дн. назад. Пора обновить:\n\n"
            f"Напиши `/backup` — я пришлю свежий файл базы."
        )

    if should_remind:
        try:
            await bot.send_message(chat_id, text, parse_mode="Markdown")
            set_chat_setting(chat_id, last_backup_remind=today_str)
            log.info(f"Sent backup reminder to {chat_id}")
        except Exception:
            log.exception(f"Failed to send backup reminder to {chat_id}")


async def _maybe_remind_pay(chat_id: int, settings, now_local, today_str: str):
    """Напоминание об оплате Railway. За 7, 3, 1 день и в день оплаты."""
    pay_date = settings["pay_date"] if "pay_date" in settings.keys() else None
    if not pay_date:
        return
    try:
        pd = datetime.strptime(pay_date, "%Y-%m-%d")
        pd_date = pd.date()
    except ValueError:
        return

    today = now_local.date()
    days_left = (pd_date - today).days

    # Шлю в дни 7, 3, 1, 0, и если просрочено (-1, -2 ... до -7)
    remind_at_days = (7, 3, 1, 0)
    if days_left not in remind_at_days and not (-7 <= days_left < 0):
        return

    last_remind = settings["last_pay_remind"] if "last_pay_remind" in settings.keys() else None
    if last_remind == today_str:
        return

    if days_left > 0:
        text = (
            f"💳 *Напоминание об оплате Railway*\n\n"
            f"До окончания триала / оплаты осталось *{days_left} {'день' if days_left == 1 else 'дн.'}* "
            f"({pd.strftime('%d.%m.%Y')}).\n\n"
            "Чтобы бот не выключился — зайди на railway.com и привяжи карту "
            "(или пополни баланс). Стоимость ≈ $5/мес.\n\n"
            "После оплаты задай новую дату: `/setpaydate ДД.ММ.ГГГГ`"
        )
    elif days_left == 0:
        text = (
            "💳 *СЕГОДНЯ дата оплаты Railway!*\n\n"
            "Если не оплатил — бот может выключиться в любой момент. "
            "Срочно зайди на railway.com.\n\n"
            "После оплаты задай новую дату: `/setpaydate ДД.ММ.ГГГГ`"
        )
    else:
        text = (
            f"⚠️ *Оплата Railway просрочена на {-days_left} дн.*\n\n"
            "Если бот ещё работает — повезло, но в любой момент может встать. "
            "Срочно оплати и задай новую дату: `/setpaydate ДД.ММ.ГГГГ`"
        )

    try:
        await bot.send_message(chat_id, text, parse_mode="Markdown")
        set_chat_setting(chat_id, last_pay_remind=today_str)
        log.info(f"Sent pay reminder to {chat_id} (days_left={days_left})")
    except Exception:
        log.exception(f"Failed to send pay reminder to {chat_id}")


async def main():
    init_db()
    log.info("Бот запущен. Жду сообщения…")
    # Фоновая задача напоминаний
    asyncio.create_task(reminders_loop())
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
